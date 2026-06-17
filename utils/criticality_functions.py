"""
Shared functions for the 5a (main-network hazard criticality) and
5c (flood-scenario accessibility plotting) analysis steps.

Notebooks call these with hardcoded paths; scripts call them via NetworkConfig
attributes.  All saved vector outputs are reprojected to EPSG:6316 and written
to Parquet, an ArcGIS File Geodatabase layer, and an Excel attribute table.
"""

from __future__ import annotations

import pickle
import string
from pathlib import Path
from typing import Any

import contextily as cx
import geopandas as gpd
import igraph as ig
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import rasterio
import xarray as xr
from damagescanner.core import DamageScanner
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from rasterstats import zonal_stats
from scipy import stats as scipy_stats
from tqdm import tqdm

from utils.arcgis import save_lyrx_layer
from utils.hazard_functions import (
    _grid_figsize,
    _set_common_extent,
    _show_or_close,
    save_excel_mirror,
    save_hazard_vector,
)

tqdm.pandas()


# ---------------------------------------------------------------------------
# Generic save helper: parquet + GDB layer + Excel attribute table (EPSG:6316)
# ---------------------------------------------------------------------------

def save_criticality_vector(
    gdf: gpd.GeoDataFrame,
    parquet_dir: Path,
    gdb_path: Path,
    layer_name: str,
    output_crs: str = "EPSG:6316",
) -> None:
    """Save *gdf* as parquet + GDB layer + mirrored Excel attribute table.

    Thin wrapper over ``save_hazard_vector``, which also writes the companion
    ``.xlsx`` under ``intermediate_results/excel/``.
    """
    save_hazard_vector(gdf, parquet_dir, gdb_path, layer_name, output_crs)


# ===========================================================================
# 5a — Main network hazard criticality
# ===========================================================================

def merge_baseline_roads(
    gdf_results: gpd.GeoDataFrame,
    baseline_roads_path: Path,
    world_path: Path,
) -> gpd.GeoDataFrame:
    """Merge criticality results with the baseline Deonice road sections.

    Sections present in the baseline network but missing from the criticality
    results are appended with zero-filled metric columns so that hazard
    exposure is evaluated for the full network. Only sections within core
    Serbia (excluding Kosovo) are considered.
    """
    gdf_deonice = gpd.read_file(baseline_roads_path)

    world = gpd.read_file(world_path)
    serbia = world.loc[world.SOV_A3 == "SRB"]
    kosovo = world.loc[world.SOV_A3 == "KOS"]
    serbia_only = gpd.overlay(serbia, kosovo, how="difference").to_crs(gdf_deonice.crs)

    gdf_deonice = gpd.sjoin(
        gdf_deonice, serbia_only[["geometry"]], how="inner", predicate="within"
    ).drop(columns=["index_right"])

    existing_ids = set(gdf_results["oznaka_deo"])
    mask_new = ~gdf_deonice["oznaka_deo"].isin(existing_ids)
    gdf_new_rows = gdf_deonice[mask_new].copy()

    shared_cols = [col for col in gdf_results.columns if col in gdf_new_rows.columns]
    cols_only_in_results = [col for col in gdf_results.columns if col not in gdf_new_rows.columns]

    gdf_new_rows = gdf_new_rows[shared_cols].copy()
    for col in cols_only_in_results:
        if gdf_results[col].dtype == object:
            gdf_new_rows[col] = ""
        else:
            gdf_new_rows[col] = 0
    gdf_new_rows = gdf_new_rows[gdf_results.columns]

    gdf_merged = pd.concat([gdf_results, gdf_new_rows], ignore_index=True)
    gdf_merged = gpd.GeoDataFrame(gdf_merged, geometry="geometry", crs=gdf_results.crs)

    print("=== Merge Summary ===")
    print(f"Original gdf_results rows:          {len(gdf_results)}")
    print(f"Deonice rows after Serbia filter:   {len(gdf_deonice)}")
    print(f"Deonice rows already in results:    {(~mask_new).sum()}  (skipped)")
    print(f"New rows added from Deonice:        {len(gdf_new_rows)}")
    print(f"Total rows in merged GeoDataFrame:  {len(gdf_merged)}")

    return gdf_merged


def calculate_flood_exposure(
    gdf_results: gpd.GeoDataFrame,
    flood_path: Path,
    world_path: Path,
) -> tuple[gpd.GeoDataFrame, gpd.GeoDataFrame]:
    """Run DamageScanner flood exposure for the road network.

    Returns (exposed_roads, country_plot). ``exposed_roads`` carries
    ``exposed`` (any depth > 0.25 m) and ``max_depth`` columns.
    """
    world = gpd.read_file(world_path)
    country_plot = world.loc[world.SOV_A3 == "SRB"]
    country_bounds = country_plot.bounds

    hazard_map = xr.open_dataset(flood_path, engine="rasterio")
    hazard_country = hazard_map.rio.clip_box(
        minx=country_bounds.minx.values[0],
        miny=country_bounds.miny.values[0],
        maxx=country_bounds.maxx.values[0],
        maxy=country_bounds.maxy.values[0],
    ).load()

    exposed_roads = DamageScanner(
        hazard_country, gdf_results, curves=pd.DataFrame(), maxdam=pd.DataFrame()
    ).exposure(asset_type="roads", disable_progress=False, return_full=False)

    exposed_roads["exposed"] = exposed_roads.progress_apply(
        lambda row: any(val > 0.25 for val in row["values"]), axis=1
    )
    exposed_roads["max_depth"] = exposed_roads.progress_apply(
        lambda row: np.max(row["values"]), axis=1
    )

    return exposed_roads, country_plot


def fill_missing_categories(exposed_roads: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Fill missing ``kategorija`` from the road designation (``oznaka_put``)."""
    cat_map = {"A1": "IA", "A2": "IA", "A3": "IA", "A4": "IA", "39": "IB"}
    mask = exposed_roads["kategorija"].isna()
    exposed_roads.loc[mask, "kategorija"] = exposed_roads.loc[mask, "oznaka_put"].map(cat_map)
    return exposed_roads


# ---------------------------------------------------------------------------
# 5a — Elevation bias analysis (road Z-profiles vs DEM)
# ---------------------------------------------------------------------------

def load_vertical_coordinates(vertical_coordinates_path: Path) -> tuple[gpd.GeoDataFrame, gpd.GeoDataFrame]:
    """Load road sections with vertical (Z) coordinates.

    Keeps only roads where every vertex has a non-zero Z value; adds
    z_min/z_max/z_range columns. Returns (metric_gdf, gdf_4326).
    """
    vertical_coordinates = gpd.read_file(vertical_coordinates_path)

    def has_complete_z(geom):
        if geom.geom_type == "MultiLineString":
            z = [c[2] for line in geom.geoms for c in line.coords]
        else:
            z = [c[2] for c in geom.coords]
        return all(v != 0 for v in z) and len(z) > 1

    vertical_coordinates = vertical_coordinates[
        vertical_coordinates.geometry.apply(has_complete_z)
    ].copy()

    vertical_coordinates["z_min"] = vertical_coordinates.geometry.apply(
        lambda g: min(c[2] for line in g.geoms for c in line.coords)
        if g.geom_type == "MultiLineString" else min(c[2] for c in g.coords)
    )
    vertical_coordinates["z_max"] = vertical_coordinates.geometry.apply(
        lambda g: max(c[2] for line in g.geoms for c in line.coords)
        if g.geom_type == "MultiLineString" else max(c[2] for c in g.coords)
    )
    vertical_coordinates["z_range"] = vertical_coordinates["z_max"] - vertical_coordinates["z_min"]

    vertical_coordinates_h = vertical_coordinates.set_crs(epsg=3909, allow_override=True)
    vertical_coordinates_4326 = vertical_coordinates_h.to_crs(epsg=4326)

    print(f"Roads with complete Z profiles: {len(vertical_coordinates)}")
    return vertical_coordinates, vertical_coordinates_4326


def extract_profiles(
    gdf_metric: gpd.GeoDataFrame,
    gdf_4326: gpd.GeoDataFrame,
    dem_path: Path,
) -> gpd.GeoDataFrame:
    """Extract per-vertex distance, measured Z, and DEM elevation profiles."""
    all_distances, all_z_values, all_dem_values = [], [], []

    with rasterio.open(dem_path) as src:
        dem_data = src.read(1)

        for idx in tqdm(range(len(gdf_metric)), desc="Extracting profiles"):
            geom_m = gdf_metric.geometry.iloc[idx]
            if geom_m.geom_type == "MultiLineString":
                coords_m = [c for line in geom_m.geoms for c in line.coords]
            else:
                coords_m = list(geom_m.coords)

            distances = [0]
            for i in range(1, len(coords_m)):
                dx = coords_m[i][0] - coords_m[i - 1][0]
                dy = coords_m[i][1] - coords_m[i - 1][1]
                distances.append(distances[-1] + (dx**2 + dy**2) ** 0.5)
            distances = [d / 1000 for d in distances]

            z_values = [c[2] if c[2] != 0 else np.nan for c in coords_m]

            geom_4326 = gdf_4326.geometry.iloc[idx]
            if geom_4326.geom_type == "MultiLineString":
                coords_4326 = [c for line in geom_4326.geoms for c in line.coords]
            else:
                coords_4326 = list(geom_4326.coords)

            dem_values = []
            for c in coords_4326:
                try:
                    row, col = src.index(c[0], c[1])
                    val = dem_data[row, col]
                    dem_values.append(val / 1000 if val != src.nodata else np.nan)
                except (IndexError, ValueError):
                    dem_values.append(np.nan)

            all_distances.append(distances)
            all_z_values.append(z_values)
            all_dem_values.append(dem_values)

    gdf_metric["distances_km"] = all_distances
    gdf_metric["z_profile"] = all_z_values
    gdf_metric["dem_profile"] = all_dem_values
    return gdf_metric


def plot_elevation_profiles(
    vertical_coordinates: gpd.GeoDataFrame,
    figure_path: Path,
    examples: dict[str, int] | None = None,
    dpi: int = 300,
    show_figures: bool = True,
) -> None:
    """Plot example measured-Z vs DEM elevation profiles, one panel per category."""
    if examples is None:
        examples = {"IA": 188, "IB": 1832, "IIA": 1881, "IIB": 1578, "IM": 453}

    fig, axes = plt.subplots(len(examples), 1, figsize=(12, 22))
    panels = [f"{letter}." for letter in string.ascii_uppercase]

    for i, (cat, idx) in enumerate(examples.items()):
        row = vertical_coordinates.loc[idx]

        axes[i].plot(row["distances_km"], row["z_profile"], label="Measured Z", linewidth=2)
        axes[i].plot(row["distances_km"], row["dem_profile"], label="DEM Elevation",
                     linewidth=2, linestyle="--", color="brown")
        axes[i].fill_between(
            row["distances_km"], row["z_profile"], row["dem_profile"],
            alpha=0.15, color="blue",
            where=[not (np.isnan(z) or np.isnan(d))
                   for z, d in zip(row["z_profile"], row["dem_profile"])],
        )

        title = f"{panels[i]} {row['PutOzn']} ({row['PutKateg']}) — {row['CvorPocNaz']} → {row['CvorZavNaz']}"
        axes[i].set_title(title, fontsize=13, fontweight="bold")
        axes[i].set_ylabel("Elevation (m)", fontsize=12, fontweight="bold")
        axes[i].legend(fontsize=11, loc="upper right")
        axes[i].grid(True, alpha=0.3)
        axes[i].tick_params(labelsize=11)

    axes[-1].set_xlabel("Distance along segment (km)", fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(Path(figure_path) / "elevation_profiles_by_category.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


def compute_bias_statistics(vertical_coordinates: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Compute per-road bias statistics (measured Z minus DEM) and terrain stats.

    Returns roads with > 5 valid Z vertices and mean/median/std difference plus
    DEM terrain characteristics, filtered to |mean_diff| < 20 m ("clean" set).
    """
    has_z = vertical_coordinates["z_profile"].apply(
        lambda z: sum(1 for v in z if not np.isnan(v)) > 5
    )
    subset = vertical_coordinates[has_z].copy()
    print(f"Roads with Z data: {has_z.sum()}")

    subset["mean_diff"] = subset.apply(
        lambda row: np.nanmean([z - d for z, d in zip(row["z_profile"], row["dem_profile"])
                                if not (np.isnan(z) or np.isnan(d))]), axis=1)
    subset["median_diff"] = subset.apply(
        lambda row: np.nanmedian([z - d for z, d in zip(row["z_profile"], row["dem_profile"])
                                  if not (np.isnan(z) or np.isnan(d))]), axis=1)
    subset["std_diff"] = subset.apply(
        lambda row: np.nanstd([z - d for z, d in zip(row["z_profile"], row["dem_profile"])
                               if not (np.isnan(z) or np.isnan(d))]), axis=1)

    subset["dem_min"] = subset["dem_profile"].apply(lambda d: np.nanmin(d))
    subset["dem_max"] = subset["dem_profile"].apply(lambda d: np.nanmax(d))
    subset["dem_range"] = subset["dem_max"] - subset["dem_min"]
    subset["dem_mean"] = subset["dem_profile"].apply(lambda d: np.nanmean(d))
    subset["road_length_km"] = subset["distances_km"].apply(lambda d: max(d))
    subset["dem_slope"] = subset["dem_range"] / subset["road_length_km"]

    clean = subset[subset["mean_diff"].abs() < 20]
    bias_lookup = clean.groupby("PutKateg")["mean_diff"].median().to_dict()
    print(f"Median bias per category: {bias_lookup}")

    return clean


def compute_bias_confidence_intervals(clean: gpd.GeoDataFrame) -> pd.DataFrame:
    """Bootstrap 95% CIs for median and parametric CIs for mean bias per category.

    Also prints the pooled IIA+IIB bootstrap estimate used to justify a shared
    bias value for lower-category roads.
    """
    ci_results = []
    for cat, group in clean.groupby("PutKateg"):
        data = group["mean_diff"].dropna()
        n = len(data)
        mean, std = data.mean(), data.std()

        ci_mean = scipy_stats.t.interval(0.95, df=n - 1, loc=mean, scale=std / np.sqrt(n))

        np.random.seed(42)
        boot_medians = [np.median(np.random.choice(data.values, size=n, replace=True))
                        for _ in range(10000)]
        ci_median = np.percentile(boot_medians, [2.5, 97.5])

        ci_results.append({
            "PutKateg": cat, "n": n, "median": data.median(),
            "median_ci_low": ci_median[0], "median_ci_high": ci_median[1],
            "mean": mean, "mean_ci_low": ci_mean[0], "mean_ci_high": ci_mean[1],
        })

    ci_df = pd.DataFrame(ci_results)
    print(ci_df.to_string(index=False))

    # Pooled IIA + IIB estimate
    pooled = clean[clean["PutKateg"].isin(["IIA", "IIB"])]
    n = len(pooled)
    np.random.seed(42)
    boot = [np.median(np.random.choice(pooled["mean_diff"].values, size=n, replace=True))
            for _ in range(10000)]
    ci = np.percentile(boot, [2.5, 97.5])
    print(f"\nPooled IIA+IIB (n={n})")
    print(f"Median: {pooled['mean_diff'].median():.2f} m")
    print(f"95% CI: [{ci[0]:.2f}, {ci[1]:.2f}]")

    return ci_df


def plot_elevation_bias(ci_df: pd.DataFrame, figure_path: Path, dpi: int = 300, show_figures: bool = True) -> None:
    """Plot median/mean elevation bias with 95% CIs per road category."""
    category_order = ["IM", "IA", "IB", "IIA", "IIB"]
    ci_df = ci_df.set_index("PutKateg").loc[category_order].reset_index()

    fig, ax = plt.subplots(figsize=(10, 6))
    x = range(len(ci_df))
    labels = [f"{row['PutKateg']}\n(n={row['n']})" for _, row in ci_df.iterrows()]
    ax.errorbar(x, ci_df["median"],
                yerr=[ci_df["median"] - ci_df["median_ci_low"],
                      ci_df["median_ci_high"] - ci_df["median"]],
                fmt="o", capsize=8, capthick=2, markersize=10, label="Median + 95% CI", zorder=3)
    ax.errorbar([i + 0.15 for i in x], ci_df["mean"],
                yerr=[ci_df["mean"] - ci_df["mean_ci_low"],
                      ci_df["mean_ci_high"] - ci_df["mean"]],
                fmt="s", capsize=8, capthick=2, markersize=10, color="red", label="Mean + 95% CI", zorder=3)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=13)
    ax.set_xlabel("Road Category", fontsize=14, fontweight="bold")
    ax.set_ylabel("Bias: Road - DEM (m)", fontsize=14, fontweight="bold")
    ax.axhline(y=0, color="grey", linestyle="--", alpha=0.5)
    ax.legend(fontsize=12)
    ax.tick_params(axis="y", labelsize=12)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(Path(figure_path) / "elevation_bias_by_category.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


# Final elevation bias (m) per road category derived from the bootstrap analysis
FINAL_ELEVATION_BIAS = {"IA": 4.601, "IM": 3.659, "IB": 0.82, "IIA": 0.82, "IIB": 0.82}


def apply_bias_correction(
    exposed_roads: gpd.GeoDataFrame,
    final_bias: dict[str, float] | None = None,
) -> gpd.GeoDataFrame:
    """Subtract the per-category elevation bias from max flood depth (clipped at 0)."""
    if final_bias is None:
        final_bias = FINAL_ELEVATION_BIAS

    exposed_roads["bias"] = exposed_roads["kategorija"].map(final_bias)
    exposed_roads["corrected_max_depth"] = (
        exposed_roads["max_depth"] - exposed_roads["bias"]
    ).clip(lower=0)

    summary = exposed_roads.groupby("kategorija").agg(
        total=("max_depth", "count"),
        originally_flooded=("max_depth", lambda x: (x > 0).sum()),
        still_flooded=("corrected_max_depth", lambda x: (x > 0).sum()),
        bias=("bias", "first"),
        avg_original_depth=("max_depth", "mean"),
        avg_corrected_depth=("corrected_max_depth", "mean"),
    ).reset_index()
    summary["removed"] = summary["originally_flooded"] - summary["still_flooded"]
    summary["pct_removed"] = (summary["removed"] / summary["originally_flooded"] * 100).round(1)
    print(summary.to_string(index=False))

    return exposed_roads


def plot_flood_exposure_correction(
    exposed_roads: gpd.GeoDataFrame,
    country_plot: gpd.GeoDataFrame,
    figure_path: Path,
    dpi: int = 300,
    show_figures: bool = True,
) -> None:
    """Two-panel map: flood exposure before (A) and after (B) bias correction."""
    gdf_mercator = exposed_roads.to_crs(3857)
    bounds_3857 = gdf_mercator.total_bounds

    figsize = _grid_figsize(bounds_3857, n_rows=1, n_cols=2, panel_height=8.0)
    fig, axes = plt.subplots(1, 2, figsize=figsize, facecolor="white")
    fig.subplots_adjust(left=0.0, right=1.0, top=1.0, bottom=0.0, wspace=0.02)

    for idx, (label, col) in enumerate([("A", "max_depth"), ("B", "corrected_max_depth")]):
        ax = axes[idx]

        not_flooded = gdf_mercator[gdf_mercator[col] == 0]
        if not not_flooded.empty:
            not_flooded.plot(ax=ax, color="#c1121f", linewidth=1.5, zorder=2)
        flooded = gdf_mercator[gdf_mercator[col] > 0]
        if not flooded.empty:
            flooded.plot(ax=ax, color="#2171b5", linewidth=1.5, zorder=3)

        ax.text(0.05, 0.95, label, transform=ax.transAxes, fontsize=20, fontweight="bold",
                verticalalignment="top", zorder=10,
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))

    _set_common_extent(axes, bounds_3857)
    for ax in axes:
        cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, attribution=False)
        ax.set_aspect("equal")
        ax.axis("off")

    legend_elements = [
        Line2D([0], [0], color="#2171b5", lw=2, label="Flooded"),
        Line2D([0], [0], color="#c1121f", lw=1, label="Not flooded"),
    ]
    axes[1].legend(handles=legend_elements, loc="lower right", fontsize=12,
                   frameon=True, fancybox=True, framealpha=0.9,
                   facecolor="white", edgecolor="#cccccc")

    plt.savefig(Path(figure_path) / "flood_exposure_correction.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


def plot_vhl_flooded_map(
    gdf_vhl_flooded: gpd.GeoDataFrame,
    figure_path: Path,
    dpi: int = 300,
    show_figures: bool = True,
) -> None:
    """Map of vehicle-hours-lost for flood-exposed roads (bias-corrected)."""
    bins = [0, 1000, 5000, 10000, 25000, np.inf]
    labels = ["0-1K", "1K-5K", "5K-10K", "10K-25K", "25K+"]
    gdf_vhl_flooded["vhl_class"] = pd.cut(gdf_vhl_flooded["vhl"], bins=bins, labels=labels, include_lowest=True)

    linewidth_map = {"0-1K": 0.5, "1K-5K": 1.0, "5K-10K": 2.0, "10K-25K": 3.5, "25K+": 5.0}
    colors = ["#fee5d9", "#fcae91", "#fb6a4a", "#de2d26", "#a50f15"]

    fig, ax = plt.subplots(1, 1, figsize=(20, 8), facecolor="white")
    for i, (class_name, width) in enumerate(linewidth_map.items()):
        subset = gdf_vhl_flooded[gdf_vhl_flooded["vhl_class"] == class_name]
        if not subset.empty:
            subset.to_crs(3857).plot(ax=ax, color=colors[i], linewidth=width, alpha=0.8, label=class_name)

    cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, attribution=False)
    ax.set_aspect("equal")
    ax.axis("off")

    legend_elements = [Line2D([0], [0], color=colors[i], lw=width, label=f"{class_name} vehicle hours")
                       for i, (class_name, width) in enumerate(linewidth_map.items())]
    ax.legend(handles=legend_elements, title="Vehicle Hours Lost", loc="upper right",
              fontsize=9, title_fontsize=10, frameon=True, fancybox=True, shadow=True,
              framealpha=0.9, facecolor="white", edgecolor="#cccccc")

    plt.tight_layout()
    plt.subplots_adjust(top=0.88, bottom=0.08, left=0.02, right=0.94)
    plt.savefig(Path(figure_path) / "vehicle_hours_lost_map_flooded.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


# ---------------------------------------------------------------------------
# 5a — Other hazard overlays
# ---------------------------------------------------------------------------

# Attribute columns carried over from the criticality results into the
# per-hazard overlays
_OVERLAY_COLS = [
    "from_id", "to_id", "objectid", "oznaka_deo", "smer_gdf1", "kategorija",
    "oznaka_put", "oznaka_poc", "naziv_poce", "oznaka_zav", "naziv_zavr",
    "duzina_deo", "pocetna_st", "zavrsna_st", "stanje", "geometry", "id",
    "passenger_cars", "buses", "light_trucks", "medium_trucks",
    "heavy_trucks", "articulated_vehicles", "total_aadt", "road_length",
    "speed", "fft", "edge_no", "vhl", "phl", "thl", "pkl", "tkl",
]


def calculate_vhl_snowdrift(
    gdf_results: gpd.GeoDataFrame, snow_drift_path: Path
) -> gpd.GeoDataFrame:
    """Spatially join road segments with snow-drift locations."""
    snow_drift = gpd.read_file(snow_drift_path)
    return gdf_results[_OVERLAY_COLS].sjoin(snow_drift)


def calculate_vhl_landslides(
    gdf_results: gpd.GeoDataFrame, landslide_path: Path
) -> gpd.GeoDataFrame:
    """Spatially join road segments with landslides (tip=='Klizište', 10 m buffer)."""
    landslides = gpd.read_file(landslide_path)
    landslides = landslides[landslides["tip"] == "Klizište"]
    landslides.geometry = landslides.geometry.buffer(10)
    return gdf_results[_OVERLAY_COLS].sjoin(landslides)


def calculate_wildfire_exposure(
    gdf_results: gpd.GeoDataFrame,
    wildfire_path: Path,
    world_path: Path,
) -> gpd.GeoDataFrame:
    """Binary wildfire exposure via DamageScanner (any susceptibility class > 0).

    Adds ``wildfire_susc`` (=1 for all exposed segments) plus coverage columns.
    """
    wildfire_map = xr.open_dataset(wildfire_path, engine="rasterio")

    wildfire_binary = wildfire_map.where(wildfire_map > 0, other=np.nan)
    wildfire_binary = wildfire_binary.where(wildfire_binary.isnull(), other=1.0)

    world = gpd.read_file(world_path)
    country_bounds = world.loc[world.SOV_A3 == "SRB"].to_crs(epsg=32634).bounds

    wildfire_country = wildfire_binary.rio.clip_box(
        minx=country_bounds.minx.values[0],
        miny=country_bounds.miny.values[0],
        maxx=country_bounds.maxx.values[0],
        maxy=country_bounds.maxy.values[0],
    ).load()

    gdf_results_32634 = gdf_results.to_crs(epsg=32634)
    exposed_wildfire = DamageScanner(
        wildfire_country, gdf_results_32634, curves=pd.DataFrame(), maxdam=pd.DataFrame()
    ).exposure(asset_type="roads", disable_progress=False, return_full=False)

    exposed_wildfire["wildfire_coverage_m"] = exposed_wildfire["coverage"].apply(lambda c: sum(c))
    exposed_wildfire["wildfire_coverage_pct"] = (
        exposed_wildfire["wildfire_coverage_m"] / (exposed_wildfire["road_length"] * 1000) * 100
    )
    exposed_wildfire["wildfire_susc"] = 1

    return exposed_wildfire.to_crs(gdf_results.crs).copy()


def calculate_heat_exposure(
    gdf_results: gpd.GeoDataFrame, pavement_temperature_path: Path
) -> gpd.GeoDataFrame:
    """Max future pavement temperature (UHI raster) per road via zonal statistics."""
    gdf_results_4326 = gdf_results.to_crs("EPSG:4326")
    stats = zonal_stats(
        gdf_results_4326, str(pavement_temperature_path),
        stats=["max"], all_touched=True, nodata=np.nan,
    )
    exposed_heat = gdf_results_4326.copy()
    exposed_heat["max_pavement_temp"] = [s["max"] for s in stats]
    return exposed_heat.to_crs(gdf_results.crs).copy()


def combine_hazard_exposure(
    gdf_results: gpd.GeoDataFrame,
    gdf_vhl_flooded: gpd.GeoDataFrame,
    gdf_vhl_snowdrift: gpd.GeoDataFrame,
    gdf_vhl_landslides: gpd.GeoDataFrame,
    gdf_vhl_wildfire: gpd.GeoDataFrame,
    gdf_vhl_heat: gpd.GeoDataFrame,
) -> gpd.GeoDataFrame:
    """Combine the per-hazard overlays into one main-network exposure GeoDataFrame.

    Duplicates per base index are aggregated with max; landslide dates are
    formatted as dd/mm/yyyy strings; only segments exposed to at least one
    hazard with a valid VHL value are kept.  Mixed-type object columns are
    cast to string so the result can be written to parquet/GDB/Excel.
    """
    s_depth = gdf_vhl_flooded["corrected_max_depth"].groupby(level=0).max()
    s_snow = gdf_vhl_snowdrift["dužina_sn"].groupby(level=0).max()
    s_date = gdf_vhl_landslides["datum_evid"].groupby(level=0).max()
    s_wildfire = gdf_vhl_wildfire["wildfire_susc"].groupby(level=0).max()
    s_heat = gdf_vhl_heat["max_pavement_temp"].groupby(level=0).max()

    gdf_hazards = pd.concat(
        [gdf_results, s_depth.rename("max_depth"), s_snow.rename("dužina_sn"),
         s_date.rename("datum_evid"), s_wildfire, s_heat],
        axis=1,
    )
    gdf_hazards["datum_evid"] = gdf_hazards["datum_evid"].dt.strftime("%d/%m/%Y")

    keep_attrs = [
        "oznaka_deo", "smer_gdf1", "kategorija", "oznaka_put", "oznaka_poc",
        "naziv_poce", "oznaka_zav", "naziv_zavr", "duzina_deo", "pocetna_st",
        "zavrsna_st", "stanje", "geometry", "passenger_cars", "buses",
        "light_trucks", "medium_trucks", "heavy_trucks", "articulated_vehicles",
        "total_aadt", "road_length", "average_time_disruption",
        "vhl", "phl", "thl", "pkl", "tkl",
        "max_depth", "dužina_sn", "datum_evid", "wildfire_susc", "max_pavement_temp",
    ]
    gdf_hazards = gdf_hazards[keep_attrs]
    gdf_hazards = gdf_hazards.loc[
        gdf_hazards[["max_depth", "dužina_sn", "datum_evid", "wildfire_susc", "max_pavement_temp"]].any(axis=1)
    ]
    gdf_hazards = gdf_hazards.loc[gdf_hazards["vhl"].notna()]

    # Normalize mixed-type object columns so parquet/GDB/Excel writers accept them
    for col in gdf_hazards.select_dtypes(include="object").columns:
        if col == gdf_hazards.geometry.name:
            continue
        has_non_string = gdf_hazards[col].dropna().apply(lambda x: not isinstance(x, str)).any()
        if has_non_string:
            print(f"  Mixed types in '{col}' (dtype=object) — casting to string")
            gdf_hazards[col] = gdf_hazards[col].astype(str)

    return gdf_hazards


def plot_hazard_comparison_grid(
    datasets: dict[str, tuple[str, gpd.GeoDataFrame]],
    column: str,
    bins: list,
    labels: list[str],
    linewidth_map: dict[str, float],
    legend_title: str,
    legend_unit: str,
    country_plot: gpd.GeoDataFrame,
    figure_path: Path,
    file_name: str,
    basemap_alpha: float = 1.0,
    dpi: int = 300,
    show_figures: bool = True,
) -> None:
    """2×2 hazard-comparison maps (A–D) for a disruption metric + legend strip.

    ``datasets`` maps panel letter -> (title, GeoDataFrame); the metric in
    *column* is binned into *labels* and styled via *linewidth_map*.
    """
    # Lightest tint (former lowest category) omitted: the four categories now
    # use the upper four colours so the darkest/thickest styling stays on the
    # open-ended top category.
    colors = ["#fcae91", "#fb6a4a", "#de2d26", "#a50f15"]
    class_col = f"{column}_class"

    for _, (_, gdf) in datasets.items():
        gdf[class_col] = pd.cut(gdf[column], bins=bins, labels=labels, include_lowest=True)

    bounds_3857 = country_plot.to_crs(3857).total_bounds
    figw, figh = _grid_figsize(bounds_3857, n_rows=2, n_cols=2, panel_height=6.5)

    fig = plt.figure(figsize=(figw, figh + 1.0), facecolor="white")
    gs = fig.add_gridspec(3, 4, height_ratios=[1, 1, 0.05], hspace=0.08, wspace=0.04,
                          left=0.01, right=0.99, top=0.96, bottom=0.02)
    axes_map = {
        "A": fig.add_subplot(gs[0, 0:2]), "B": fig.add_subplot(gs[0, 2:4]),
        "C": fig.add_subplot(gs[1, 0:2]), "D": fig.add_subplot(gs[1, 2:4]),
    }
    ax_legend = fig.add_subplot(gs[2, :])

    for letter, (title, gdf) in datasets.items():
        ax = axes_map[letter]
        gdf_mercator = gdf.to_crs(3857)

        for i, (class_name, width) in enumerate(linewidth_map.items()):
            subset = gdf_mercator[gdf_mercator[class_col] == class_name]
            if not subset.empty:
                subset.plot(ax=ax, color=colors[i], linewidth=width, zorder=2)

        ax.set_title(title, fontsize=14, fontweight="bold", pad=6)
        ax.text(0.05, 0.95, letter, transform=ax.transAxes, fontsize=20, fontweight="bold",
                verticalalignment="top", zorder=10,
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))

    _set_common_extent(list(axes_map.values()), bounds_3857)
    for ax in axes_map.values():
        cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, alpha=basemap_alpha, attribution=False)
        ax.set_aspect("equal")
        ax.axis("off")

    ax_legend.axis("off")
    legend_elements = [
        Line2D([0], [0], color=colors[i], lw=width * 1.5, label=f"{class_name} {legend_unit}")
        for i, (class_name, width) in enumerate(linewidth_map.items())
    ]
    ax_legend.legend(handles=legend_elements, title=legend_title, loc="center",
                     fontsize=10, title_fontsize=11, frameon=True, fancybox=True,
                     shadow=True, framealpha=0.9, facecolor="white", ncols=5,
                     edgecolor="#cccccc")

    plt.savefig(Path(figure_path) / file_name, dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


def plot_all_hazard_comparisons(
    gdf_vhl_flooded: gpd.GeoDataFrame,
    gdf_vhl_snowdrift: gpd.GeoDataFrame,
    gdf_vhl_landslides: gpd.GeoDataFrame,
    gdf_vhl_wildfire: gpd.GeoDataFrame,
    country_plot: gpd.GeoDataFrame,
    figure_path: Path,
    dpi: int = 300,
    show_figures: bool = True,
) -> None:
    """VHL / PHL / TKL 2×2 hazard-comparison figures (Floods, Snow, Landslides, Wildfires)."""
    datasets = {
        "A": ("Floods", gdf_vhl_flooded),
        "B": ("Snow Drift", gdf_vhl_snowdrift),
        "C": ("Landslides", gdf_vhl_landslides),
        "D": ("Wildfires", gdf_vhl_wildfire),
    }

    plot_hazard_comparison_grid(
        datasets, column="vhl",
        bins=[0, 1000, 5000, 10000, np.inf],
        labels=["0-1K", "1K-5K", "5K-10K", "10K+"],
        linewidth_map={"0-1K": 1.5, "1K-5K": 2.0, "5K-10K": 3.5, "10K+": 5.0},
        legend_title="Vehicle Hours Lost", legend_unit="vehicle hours",
        country_plot=country_plot, figure_path=figure_path,
        file_name="vhl_hazards_comparison.png", basemap_alpha=0.4, dpi=dpi,
        show_figures=show_figures,
    )

    plot_hazard_comparison_grid(
        datasets, column="phl",
        bins=[0, 1000, 5000, 10000, np.inf],
        labels=["0-1K", "1K-5K", "5K-10K", "10K+"],
        linewidth_map={"0-1K": 1.5, "1K-5K": 2.0, "5K-10K": 2.5, "10K+": 3.0},
        legend_title="Passenger Hours Lost", legend_unit="hours",
        country_plot=country_plot, figure_path=figure_path,
        file_name="phl_hazards_comparison.png", basemap_alpha=1.0, dpi=dpi,
        show_figures=show_figures,
    )

    plot_hazard_comparison_grid(
        datasets, column="tkl",
        bins=[10000, 25000, 50000, 100000, np.inf],
        labels=["10K-25K", "25K-50K", "50K-100K", "100K+"],
        linewidth_map={"10K-25K": 1.5, "25K-50K": 2.0, "50K-100K": 2.5, "100K+": 3.0},
        legend_title="Tonnage Kilometers Lost", legend_unit="ton-km",
        country_plot=country_plot, figure_path=figure_path,
        file_name="tkl_hazards_comparison.png", basemap_alpha=1.0, dpi=dpi,
        show_figures=show_figures,
    )


def print_hazard_analysis_summary(
    gdf_results: gpd.GeoDataFrame,
    gdf_vhl_flooded: gpd.GeoDataFrame,
    gdf_vhl_snowdrift: gpd.GeoDataFrame,
    gdf_vhl_landslides: gpd.GeoDataFrame,
    gdf_vhl_wildfire: gpd.GeoDataFrame,
) -> None:
    """Print disruption-metric summaries, overlap analysis, and national comparison."""
    print("=" * 70)
    print("BASELINE RESULTS - ALL METRICS")
    print("=" * 70)
    print(f"\nTotal road segments analyzed: {len(gdf_results):,}")
    print("\nAll Metrics Summary:")
    print(gdf_results[["phl", "thl", "pkl", "tkl"]].describe().round(2))
    print("\nTotals across network:")
    print(f"  Total PHL: {gdf_results['phl'].sum():,.0f} passenger hours")
    print(f"  Total THL: {gdf_results['thl'].sum():,.0f} ton hours")
    print(f"  Total PKL: {gdf_results['pkl'].sum():,.0f} passenger km")
    print(f"  Total TKL: {gdf_results['tkl'].sum():,.0f} ton km")

    gdf_vhl_snowdrift = gdf_vhl_snowdrift.rename(columns={"oznaka_deo_left": "oznaka_deo"})
    gdf_vhl_landslides = gdf_vhl_landslides.rename(columns={"oznaka_deo_left": "oznaka_deo"})

    hazard_datasets = {
        "Floods": gdf_vhl_flooded,
        "Snow Drift": gdf_vhl_snowdrift,
        "Landslides": gdf_vhl_landslides,
        "Wildfires": gdf_vhl_wildfire,
    }

    bins_phl = [0, 1000, 5000, 10000, 25000, np.inf]
    labels_phl = ["0-1K", "1K-5K", "5K-10K", "10K-25K", "25K+"]
    bins_tkl = [0, 10000, 50000, 100000, 250000, np.inf]
    labels_tkl = ["0-10K", "10K-50K", "50K-100K", "100K-250K", "250K+"]

    # --- PHL ---
    print("\n" + "=" * 70)
    print("PASSENGER HOURS LOST (PHL) BY HAZARD")
    print("=" * 70)
    phl_summary = []
    for hazard_name, gdf in hazard_datasets.items():
        print(f"\n{'─' * 50}\n{hazard_name.upper()}\n{'─' * 50}")
        print(f"Exposed road segments: {len(gdf):,}")
        print(gdf["phl"].describe().round(2))
        gdf["phl_class"] = pd.cut(gdf["phl"], bins=bins_phl, labels=labels_phl, include_lowest=True)
        phl_counts = gdf["phl_class"].value_counts().sort_index()
        phl_pcts = (gdf["phl_class"].value_counts(normalize=True).sort_index() * 100).round(1)
        for label in labels_phl:
            print(f"  {label}: {phl_counts.get(label, 0)} ({phl_pcts.get(label, 0)}%)")
        if "kategorija" in gdf.columns:
            road_summary = gdf.groupby("kategorija")["phl"].agg(["count", "sum", "mean", "max"]).round(2)
            road_summary.columns = ["Count", "Total PHL", "Mean PHL", "Max PHL"]
            print(road_summary.sort_values("Total PHL", ascending=False))
        cols = [c for c in ["oznaka_deo", "kategorija", "naziv_poce", "naziv_zavr", "total_aadt", "phl"] if c in gdf.columns]
        print(gdf.nlargest(5, "phl")[cols].to_string())
        phl_summary.append({
            "Hazard": hazard_name, "Exposed Segments": len(gdf),
            "Total PHL": gdf["phl"].sum(), "Mean PHL": gdf["phl"].mean(),
            "Median PHL": gdf["phl"].median(), "Max PHL": gdf["phl"].max(),
            "Segments >25K": len(gdf[gdf["phl"] >= 25000]),
            "Segments >10K": len(gdf[gdf["phl"] >= 10000]),
        })
    print("\n" + "─" * 50 + "\nPHL HAZARD COMPARISON SUMMARY\n" + "─" * 50)
    print(pd.DataFrame(phl_summary).to_string(index=False))

    # --- TKL ---
    print("\n" + "=" * 70)
    print("TONNAGE KILOMETERS LOST (TKL) BY HAZARD")
    print("=" * 70)
    tkl_summary = []
    for hazard_name, gdf in hazard_datasets.items():
        print(f"\n{'─' * 50}\n{hazard_name.upper()}\n{'─' * 50}")
        print(f"Exposed road segments: {len(gdf):,}")
        print(gdf["tkl"].describe().round(2))
        gdf["tkl_class"] = pd.cut(gdf["tkl"], bins=bins_tkl, labels=labels_tkl, include_lowest=True)
        tkl_counts = gdf["tkl_class"].value_counts().sort_index()
        tkl_pcts = (gdf["tkl_class"].value_counts(normalize=True).sort_index() * 100).round(1)
        for label in labels_tkl:
            print(f"  {label}: {tkl_counts.get(label, 0)} ({tkl_pcts.get(label, 0)}%)")
        if "kategorija" in gdf.columns:
            road_summary = gdf.groupby("kategorija")["tkl"].agg(["count", "sum", "mean", "max"]).round(2)
            road_summary.columns = ["Count", "Total TKL", "Mean TKL", "Max TKL"]
            print(road_summary.sort_values("Total TKL", ascending=False))
        cols = [c for c in ["oznaka_deo", "kategorija", "naziv_poce", "naziv_zavr", "total_aadt", "tkl"] if c in gdf.columns]
        print(gdf.nlargest(5, "tkl")[cols].to_string())
        tkl_summary.append({
            "Hazard": hazard_name, "Exposed Segments": len(gdf),
            "Total TKL": gdf["tkl"].sum(), "Mean TKL": gdf["tkl"].mean(),
            "Median TKL": gdf["tkl"].median(), "Max TKL": gdf["tkl"].max(),
            "Segments >250K": len(gdf[gdf["tkl"] >= 250000]),
            "Segments >100K": len(gdf[gdf["tkl"] >= 100000]),
        })
    print("\n" + "─" * 50 + "\nTKL HAZARD COMPARISON SUMMARY\n" + "─" * 50)
    print(pd.DataFrame(tkl_summary).to_string(index=False))

    # --- THL & PKL ---
    for metric, label, threshold in [("thl", "TONNAGE HOURS LOST (THL)", 10000),
                                     ("pkl", "PASSENGER KILOMETERS LOST (PKL)", 250000)]:
        print("\n" + "=" * 70)
        print(f"{label} BY HAZARD")
        print("=" * 70)
        summary = []
        for hazard_name, gdf in hazard_datasets.items():
            summary.append({
                "Hazard": hazard_name, "Exposed Segments": len(gdf),
                f"Total {metric.upper()}": gdf[metric].sum(),
                f"Mean {metric.upper()}": gdf[metric].mean(),
                f"Max {metric.upper()}": gdf[metric].max(),
                f"Segments >{threshold // 1000}K": len(gdf[gdf[metric] >= threshold]),
            })
        print(pd.DataFrame(summary).to_string(index=False))

    # --- Overlap analysis ---
    print("\n" + "=" * 70)
    print("OVERLAP ANALYSIS - MULTI-HAZARD EXPOSURE")
    print("=" * 70)
    seg_sets = {name: set(gdf["oznaka_deo"].dropna()) for name, gdf in hazard_datasets.items()}
    for name, s in seg_sets.items():
        print(f"Total unique segments exposed to {name}: {len(s)}")
    all_names = list(seg_sets.keys())
    all_exposed = set().union(*seg_sets.values())
    print(f"\nTotal unique segments exposed to any hazard: {len(all_exposed)}")

    print("\nExclusive exposure:")
    for name, s in seg_sets.items():
        others = set().union(*(seg_sets[n] for n in all_names if n != name))
        print(f"  {name} only: {len(s - others)}")

    print("\nDual exposure (all pairs):")
    for i, n1 in enumerate(all_names):
        for n2 in all_names[i + 1:]:
            others = set().union(*(seg_sets[n] for n in all_names if n not in (n1, n2)))
            print(f"  {n1} AND {n2}: {len(seg_sets[n1] & seg_sets[n2] - others)}")

    print(f"\nAll four hazards: {len(set.intersection(*seg_sets.values()))}")

    from collections import Counter
    segment_hazard_count = Counter()
    for seg in all_exposed:
        n = sum(1 for s in seg_sets.values() if seg in s)
        segment_hazard_count[n] += 1
    print("\nSegments by number of hazards:")
    for n in sorted(segment_hazard_count.keys()):
        print(f"  {n} hazard(s): {segment_hazard_count[n]} segments")

    # --- National comparison ---
    print("\n" + "=" * 70)
    print("HAZARD EXPOSURE AS % OF NATIONAL DAILY TRANSPORT")
    print("=" * 70)
    national_pkm_daily = 2069 * 1e6 / 180
    national_tkm_daily = 4677 * 1e6 / 180
    print("\nNational daily averages (road transport, H1 2025):")
    print(f"  Passenger-km/day: {national_pkm_daily:,.0f}")
    print(f"  Ton-km/day: {national_tkm_daily:,.0f}")
    print("\nTotal exposed PKL as % of national daily:")
    for name, gdf in hazard_datasets.items():
        total = gdf["pkl"].sum()
        print(f"  {name}: {total:,.0f} PKL ({total / national_pkm_daily * 100:.1f}% of national daily)")
    print("\nTotal exposed TKL as % of national daily:")
    for name, gdf in hazard_datasets.items():
        total = gdf["tkl"].sum()
        print(f"  {name}: {total:,.0f} TKL ({total / national_tkm_daily * 100:.1f}% of national daily)")

    # --- Key examples ---
    print("\n" + "=" * 70)
    print("KEY EXAMPLES FOR TEXT")
    print("=" * 70)
    for hazard_name, gdf in hazard_datasets.items():
        print(f"\n{'─' * 50}\n{hazard_name.upper()} - TOP CRITICAL SEGMENTS\n{'─' * 50}")
        print("\nTop 3 by Passenger Hours Lost:")
        cols = [c for c in ["oznaka_deo", "kategorija", "oznaka_put", "naziv_poce", "naziv_zavr", "total_aadt", "phl", "pkl"] if c in gdf.columns]
        for _, row in gdf.nlargest(3, "phl")[cols].iterrows():
            print(f"  {row.get('oznaka_put', 'N/A')} ({row.get('kategorija', 'N/A')}): "
                  f"{row.get('naziv_poce', 'N/A')} → {row.get('naziv_zavr', 'N/A')}")
            print(f"    AADT: {row.get('total_aadt', 0):,.0f}, PHL: {row['phl']:,.0f}, PKL: {row.get('pkl', 0):,.0f}")
        print("\nTop 3 by Tonnage Kilometers Lost:")
        cols = [c for c in ["oznaka_deo", "kategorija", "oznaka_put", "naziv_poce", "naziv_zavr", "total_aadt", "thl", "tkl"] if c in gdf.columns]
        for _, row in gdf.nlargest(3, "tkl")[cols].iterrows():
            print(f"  {row.get('oznaka_put', 'N/A')} ({row.get('kategorija', 'N/A')}): "
                  f"{row.get('naziv_poce', 'N/A')} → {row.get('naziv_zavr', 'N/A')}")
            print(f"    AADT: {row.get('total_aadt', 0):,.0f}, THL: {row.get('thl', 0):,.0f}, TKL: {row['tkl']:,.0f}")

    # --- Master summary table ---
    print("\n" + "=" * 70)
    print("MASTER SUMMARY TABLE FOR TEXT")
    print("=" * 70)
    master = []
    for name, gdf in hazard_datasets.items():
        master.append({
            "Hazard": name, "Segments": len(gdf),
            "Total PHL": f"{gdf['phl'].sum():,.0f}", "Total THL": f"{gdf['thl'].sum():,.0f}",
            "Total PKL": f"{gdf['pkl'].sum():,.0f}", "Total TKL": f"{gdf['tkl'].sum():,.0f}",
            "Max PHL": f"{gdf['phl'].max():,.0f}", "Max TKL": f"{gdf['tkl'].max():,.0f}",
            "PHL >25K": len(gdf[gdf["phl"] >= 25000]),
            "TKL >250K": len(gdf[gdf["tkl"] >= 250000]),
        })
    print(pd.DataFrame(master).to_string(index=False))


# ---------------------------------------------------------------------------
# 5a — Multi-hazard exposure count (binary yes/no per hazard)
# ---------------------------------------------------------------------------

def flag_future_precipitation(
    gdf: gpd.GeoDataFrame,
    precip_path: Path,
    threshold: float = 10.0,
    value_col: str = "max_rx1day_pct",
):
    """Boolean exposure flag per segment for projected extreme-rainfall change.

    Loads the future-precipitation layer (4b), spatially joins the maximum
    ``value_col`` (% change) onto *gdf*, and flags segments at or above
    *threshold* percent. Returns a numpy boolean array aligned to ``gdf.index``.
    """
    precip = gpd.read_parquet(precip_path).to_crs(gdf.crs)
    joined = add_impact_column(
        gdf[["geometry"]].copy(), precip, "_precip", agg="max", col_to_focus=value_col
    )
    return (joined["_precip"].fillna(0) >= threshold).to_numpy()


def build_hazard_count(
    gdf: gpd.GeoDataFrame,
    hazard_specs: list[tuple[str, str, Any]],
    count_col: str = "hazard_count",
) -> gpd.GeoDataFrame:
    """Add one binary column per hazard plus a total *count_col*.

    *hazard_specs* is an ordered list of ``(column_name, display_label, mask)``
    where *mask* is a boolean array/Series aligned to ``gdf.index``.
    """
    gdf = gdf.copy()
    for col, _label, mask in hazard_specs:
        gdf[col] = np.asarray(mask).astype(int)
    gdf[count_col] = gdf[[col for col, _, _ in hazard_specs]].sum(axis=1)
    return gdf


def summarize_hazard_counts(
    gdf: gpd.GeoDataFrame,
    hazard_specs: list[tuple[str, str, Any]],
    count_col: str = "hazard_count",
) -> None:
    """Print roads exposed per hazard (count + % of network) and the
    distribution of roads by number of hazards."""
    n = len(gdf)
    k_max = len(hazard_specs)
    print("=" * 70)
    print("MULTI-HAZARD EXPOSURE SUMMARY")
    print("=" * 70)
    print(f"Total road segments in network: {n}\n")

    print("Roads exposed per hazard:")
    for col, label, _ in hazard_specs:
        c = int(gdf[col].sum())
        print(f"  {label:34s}: {c:6d}  ({c / n * 100:5.1f}% of network)")

    print("\nRoads by number of hazards:")
    dist = gdf[count_col].value_counts().sort_index()
    for k in range(0, k_max + 1):
        c = int(dist.get(k, 0))
        print(f"  {k} hazard(s): {c:6d}  ({c / n * 100:5.1f}% of network)")

    exposed = int((gdf[count_col] >= 1).sum())
    print(f"\nRoads exposed to at least one hazard: {exposed} ({exposed / n * 100:.1f}% of network)")


def plot_hazard_count_map(
    gdf: gpd.GeoDataFrame,
    figure_path: Path,
    count_col: str = "hazard_count",
    n_hazards: int = 6,
    dpi: int = 300,
    show_figures: bool = True,
) -> None:
    """Map every road segment coloured by how many hazards it is exposed to.

    Segments with no hazard are drawn faintly in grey; exposed segments use a
    sequential yellow→red ramp with increasing line width per hazard count.
    """
    colors = ["#ffffb2", "#fed976", "#feb24c", "#fd8d3c", "#f03b20", "#bd0026",
              "#7a0177", "#49006a"]
    widths = {1: 0.8, 2: 1.2, 3: 1.8, 4: 2.5, 5: 3.2, 6: 4.0, 7: 4.6, 8: 5.2}

    g = gdf.to_crs(3857)
    fig, ax = plt.subplots(1, 1, figsize=(20, 8), facecolor="white")

    zero = g[g[count_col] == 0]
    if not zero.empty:
        zero.plot(ax=ax, color="#d9d9d9", linewidth=0.3, alpha=0.6, zorder=1)

    for k in range(1, n_hazards + 1):
        subset = g[g[count_col] == k]
        if not subset.empty:
            subset.plot(ax=ax, color=colors[k - 1], linewidth=widths[k], alpha=0.9, zorder=2 + k)

    cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, alpha=0.4, attribution=False)
    ax.set_aspect("equal")
    ax.axis("off")

    legend_elements = [
        Line2D([0], [0], color=colors[k - 1], lw=widths[k],
               label=f"{k} hazard" + ("s" if k > 1 else ""))
        for k in range(1, n_hazards + 1)
    ]
    ax.legend(handles=legend_elements, title="Number of hazards", loc="upper right",
              fontsize=10, title_fontsize=11, frameon=True, fancybox=True, shadow=True,
              framealpha=0.9, facecolor="white", edgecolor="#cccccc")

    plt.tight_layout()
    plt.subplots_adjust(top=0.88, bottom=0.08, left=0.02, right=0.94)
    plt.savefig(Path(figure_path) / "hazard_count_map.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


# ===========================================================================
# 5c — Flood-scenario accessibility plotting
# ===========================================================================

# Common styling for travel-time impact classes
IMPACT_BINS = [0.167, 0.333, 0.5, 0.667, 1.0, np.inf]
IMPACT_LABELS = ["10-20 min", "20-30 min", "30-40 min", "40-60 min", "60+ min"]
IMPACT_COLORS = ["#fcbba1", "#fc9272", "#ef3b2c", "#cb181d", "#a50f15"]
IMPACT_WIDTHS = {"10-20 min": 1.0, "20-30 min": 1.5, "30-40 min": 2.0, "40-60 min": 2.5, "60+ min": 3.0}


def load_basins(basins_csv_path: Path, basins_shapefile_path: Path) -> gpd.GeoDataFrame:
    """Merge basin-level flood statistics with HydroBASINS geometries."""
    basins = pd.read_csv(basins_csv_path)
    all_basins = gpd.read_file(basins_shapefile_path)
    return gpd.GeoDataFrame(basins.merge(all_basins, left_on="basinID", right_on="HYBAS_ID"))


def plot_basin_water_depths(basins: gpd.GeoDataFrame, figure_path: Path, dpi: int = 300, show_figures: bool = True) -> None:
    """Map of mean water depth per basin in five depth classes."""
    bins = [0, 1, 2, 3, 4, np.inf]
    labels = ["0-1m", "1-2m", "2-3m", "3-4m", "4m+"]
    basins["depth_class"] = pd.cut(basins["mean water depth (m)"], bins=bins, labels=labels, include_lowest=True)

    colors = ["#f7fbff", "#c6dbef", "#6baed6", "#2171b5", "#08306b"]
    fig, ax = plt.subplots(1, 1, figsize=(20, 8), facecolor="white")
    basins_mercator = basins.to_crs(3857)

    for category, color in zip(labels, colors):
        category_data = basins_mercator[basins_mercator["depth_class"] == category]
        if len(category_data) > 0:
            category_data.plot(ax=ax, color=color, linewidth=0.1, edgecolor="navy", alpha=0.8, legend=False)

    cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, alpha=0.4, attribution=False)
    ax.set_aspect("equal")
    ax.axis("off")

    legend_elements = [Patch(facecolor=colors[i], label=f"{labels[i]} depth", edgecolor="navy", linewidth=0.5)
                       for i in range(len(labels))]
    ax.legend(handles=legend_elements, title="Mean Water Depth", loc="upper right",
              fontsize=10, title_fontsize=12, frameon=True, fancybox=True, shadow=True,
              framealpha=0.9, facecolor="white", edgecolor="#cccccc")

    total_basins = len(basins)
    mean_depth = basins["mean water depth (m)"].mean()
    ax.text(0.02, 0.02, f"Total Basins: {total_basins:,} | Average Depth: {mean_depth:.2f}m",
            transform=ax.transAxes, fontsize=11,
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))

    plt.tight_layout()
    plt.subplots_adjust(top=0.88, bottom=0.08, left=0.02, right=0.94)
    plt.savefig(Path(figure_path) / "basin_water_depths.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


def load_base_network(network_parquet_path: Path) -> gpd.GeoDataFrame:
    """Load the basin-scenario base network (giant component) in EPSG:3857."""
    base_network = gpd.read_parquet(network_parquet_path)

    edges = base_network.reindex(
        ["from_id", "to_id"] + [x for x in list(base_network.columns) if x not in ["from_id", "to_id"]],
        axis=1,
    )
    graph = ig.Graph.TupleList(edges.itertuples(index=False), edge_attrs=list(edges.columns)[2:], directed=True)
    graph.vs["id"] = graph.vs["name"]
    graph = graph.connected_components().giant()
    edges = edges[edges["id"].isin(graph.es["id"])]

    return base_network.to_crs(3857)


def calculate_service_criticality(
    results_path: Path, base_network: gpd.GeoDataFrame
) -> gpd.GeoDataFrame:
    """Exposed edges with mean positive travel-time impact for one service.

    Loads per-basin disruption outcomes, assigns the basin's mean positive
    delay to its removed edges, keeps exposed edges with ≥ 10 min impact, and
    bins them into ``impact_class``.
    """
    with open(results_path, "rb") as f:
        save_new_results = pickle.load(f)

    pd.options.mode.chained_assignment = None
    pd.set_option("future.no_silent_downcasting", True)

    collect_removed_edges = []
    for basin in save_new_results.keys():
        scenario_outcome = save_new_results[basin]["scenario_outcome"]
        if isinstance(scenario_outcome, pd.DataFrame):
            subset_edges = base_network.loc[
                base_network.osm_id.astype(str).isin(save_new_results[basin]["real_edges_to_remove"])
            ].copy()
            subset_edges["travel_time_impact"] = (
                scenario_outcome.loc[scenario_outcome.Delta > 0]
                .replace([np.inf, -np.inf], 1)
                .Delta.mean()
            )
            collect_removed_edges.append(subset_edges)

    exposed_edges = gpd.GeoDataFrame(pd.concat(collect_removed_edges))[
        ["osm_id", "from_id", "to_id", "highway", "exposed", "geometry", "travel_time_impact"]
    ]
    exposed_edges = (
        exposed_edges.loc[exposed_edges.exposed]
        .reset_index(drop=True)
        .set_crs(3857, allow_override=True)
    )
    exposed_edges["travel_time_impact"] = np.where(
        np.isinf(exposed_edges["travel_time_impact"]), 1, exposed_edges["travel_time_impact"]
    )
    exposed_edges = exposed_edges.to_crs(3857)
    exposed_edges = exposed_edges[exposed_edges["travel_time_impact"] >= 0.167].copy()

    exposed_edges["impact_class"] = pd.cut(
        exposed_edges["travel_time_impact"], bins=IMPACT_BINS, labels=IMPACT_LABELS, include_lowest=True
    )
    return exposed_edges


def plot_service_criticality(
    exposed_edges: gpd.GeoDataFrame,
    base_network: gpd.GeoDataFrame,
    figure_path: Path,
    file_name: str,
    dpi: int = 200,
    show_figures: bool = True,
) -> None:
    """Single map of increased travel time for one service over the base network."""
    fig, ax = plt.subplots(1, 1, figsize=(20, 8), facecolor="white")

    base_network.plot(ax=ax, linewidth=0.1, color="lightgrey", alpha=0.5)

    edges_mercator = exposed_edges.to_crs(3857)
    for i, (category, color) in enumerate(zip(IMPACT_LABELS, IMPACT_COLORS)):
        category_data = edges_mercator[edges_mercator["impact_class"] == category]
        if len(category_data) > 0:
            category_data.plot(ax=ax, color=color, linewidth=IMPACT_WIDTHS[category],
                               alpha=0.9, zorder=5 + i)

    cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, alpha=0.4, attribution=False)
    ax.set_aspect("equal")
    ax.axis("off")

    legend_elements = [Patch(facecolor=IMPACT_COLORS[i], label=f"{IMPACT_LABELS[i]} delay",
                             edgecolor="darkred", linewidth=0.5)
                       for i in range(len(IMPACT_LABELS))]
    ax.legend(handles=legend_elements, title="Increased Travel Time", loc="upper right",
              fontsize=10, title_fontsize=12, frameon=True, fancybox=True, shadow=True,
              framealpha=0.9, facecolor="white", edgecolor="#cccccc")

    plt.tight_layout()
    plt.subplots_adjust(top=0.88, bottom=0.08, left=0.02, right=0.94)
    plt.savefig(Path(figure_path) / file_name, dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)
    plt.close()


def calculate_agri_criticality(
    results_path: Path,
    base_network: gpd.GeoDataFrame,
    delta_prefix: str = "delta_avg",
) -> dict[str, gpd.GeoDataFrame | None]:
    """Agriculture criticality per sink type ('road', 'port', 'rail').

    ``delta_prefix`` selects the population delta columns:
    'delta_avg' (average over all sinks) or 'delta_nearest' (nearest sink).
    """
    with open(results_path, "rb") as f:
        save_new_results = pickle.load(f)

    pd.options.mode.chained_assignment = None
    pd.set_option("future.no_silent_downcasting", True)
    river_basins = list(save_new_results.keys())

    sink_types = ["road", "port", "rail"]
    exposed_edges_by_type: dict[str, gpd.GeoDataFrame | None] = {}

    for sink_type in sink_types:
        print(f"Processing: {sink_type}...")
        collect_removed_edges = []

        for basin in river_basins:
            result = save_new_results[basin]
            if result.get("status") == "Error" or result.get("df_population") is None:
                continue

            df_population = result["df_population"]
            delta_col = f"{delta_prefix}_{sink_type}"
            if delta_col not in df_population.columns:
                continue

            delta_values = df_population[delta_col]
            valid_deltas = delta_values[(delta_values > 0) & (~np.isinf(delta_values))]
            if len(valid_deltas) == 0:
                continue

            removed_osm_ids = result["real_edges_to_remove"]
            subset_edges = base_network.loc[
                base_network.osm_id.astype(str).isin([str(x) for x in removed_osm_ids])
            ].copy()
            if len(subset_edges) == 0:
                continue

            subset_edges["travel_time_impact"] = valid_deltas.mean()
            collect_removed_edges.append(subset_edges)

        if len(collect_removed_edges) == 0:
            print(f"  No exposed edges found for {sink_type}")
            exposed_edges_by_type[sink_type] = None
            continue

        exposed_edges = gpd.GeoDataFrame(pd.concat(collect_removed_edges))[
            ["osm_id", "from_id", "to_id", "highway", "exposed", "geometry", "travel_time_impact"]
        ]
        exposed_edges = exposed_edges.loc[exposed_edges.exposed].reset_index(drop=True)
        exposed_edges = exposed_edges.set_crs(3857, allow_override=True)
        exposed_edges["travel_time_impact"] = np.where(
            np.isinf(exposed_edges["travel_time_impact"]), 1, exposed_edges["travel_time_impact"]
        )
        exposed_edges = exposed_edges[exposed_edges["travel_time_impact"] >= 0.167].copy()
        if len(exposed_edges) == 0:
            print(f"  No edges with impact >= 10 min for {sink_type}")
            exposed_edges_by_type[sink_type] = None
            continue

        exposed_edges = exposed_edges.to_crs(3857)
        exposed_edges["impact_class"] = pd.cut(
            exposed_edges["travel_time_impact"], bins=IMPACT_BINS, labels=IMPACT_LABELS, include_lowest=True
        )
        exposed_edges_by_type[sink_type] = exposed_edges
        print(f"  {len(exposed_edges)} edges")

    return exposed_edges_by_type


def _plot_impact_panel(ax: Any, gdf: gpd.GeoDataFrame | None, letter: str,
                       base_network: gpd.GeoDataFrame, basemap_alpha: float = 1.0,
                       bounds_3857=None) -> None:
    """One panel of exposed edges by impact class over the muted base network."""
    base_network.plot(ax=ax, linewidth=0.1, color="lightgrey", alpha=0.5)

    if gdf is not None and len(gdf) > 0:
        for category, color in zip(IMPACT_LABELS, IMPACT_COLORS):
            subset = gdf[gdf["impact_class"] == category]
            if len(subset) > 0:
                subset.plot(ax=ax, color=color, linewidth=IMPACT_WIDTHS[category], alpha=0.9)

    ax.axis("off")
    ax.text(0.05, 0.95, letter, transform=ax.transAxes, fontsize=20, fontweight="bold",
            verticalalignment="top", zorder=10,
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    if bounds_3857 is not None:
        _set_common_extent(ax, bounds_3857)
    cx.add_basemap(ax=ax, source=cx.providers.CartoDB.Positron, alpha=basemap_alpha, attribution=False)
    ax.set_aspect("equal")


def plot_agri_criticality_3x1(
    exposed_edges_by_type: dict[str, gpd.GeoDataFrame | None],
    base_network: gpd.GeoDataFrame,
    figure_path: Path,
    file_name: str,
    legend_title: str,
    dpi: int = 150,
    show_figures: bool = True,
) -> None:
    """3×1 figure: agriculture criticality for road borders (A), ports (B), rail (C)."""
    bounds_3857 = base_network.total_bounds
    figw, figh = _grid_figsize(bounds_3857, n_rows=1, n_cols=3, panel_height=6.0)

    fig, axes = plt.subplots(1, 3, figsize=(figw, figh + 0.9))
    fig.subplots_adjust(left=0.0, right=1.0, top=1.0, bottom=0.9 / (figh + 0.9), wspace=0.02)

    _plot_impact_panel(axes[0], exposed_edges_by_type["road"], "A", base_network, bounds_3857=bounds_3857)
    _plot_impact_panel(axes[1], exposed_edges_by_type["port"], "B", base_network, bounds_3857=bounds_3857)
    _plot_impact_panel(axes[2], exposed_edges_by_type["rail"], "C", base_network, bounds_3857=bounds_3857)

    legend_handles = [Patch(facecolor=IMPACT_COLORS[i], label=IMPACT_LABELS[i],
                            edgecolor="darkred", linewidth=0.5)
                      for i in range(len(IMPACT_LABELS))]
    fig.legend(handles=legend_handles, title=legend_title, loc="lower center",
               bbox_to_anchor=(0.5, 0.0), ncol=len(IMPACT_LABELS), fontsize=14,
               title_fontsize=16, frameon=True, fancybox=True, framealpha=0.9)

    plt.savefig(Path(figure_path) / file_name, dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


def plot_criticality_2x2(
    hospital_exposed_edges: gpd.GeoDataFrame,
    factory_exposed_edges: gpd.GeoDataFrame,
    police_exposed_edges: gpd.GeoDataFrame,
    fire_exposed_edges: gpd.GeoDataFrame,
    base_network: gpd.GeoDataFrame,
    figure_path: Path,
    dpi: int = 150,
    show_figures: bool = True,
) -> None:
    """Combined 2×2 figure: hospitals (A), factories (B), police (C), fire (D)."""
    bounds_3857 = base_network.total_bounds
    figw, figh = _grid_figsize(bounds_3857, n_rows=2, n_cols=2, panel_height=6.5)

    fig, axes = plt.subplots(2, 2, figsize=(figw, figh + 0.9))
    fig.subplots_adjust(left=0.0, right=1.0, top=1.0, bottom=0.9 / (figh + 0.9),
                        wspace=0.02, hspace=0.02)

    _plot_impact_panel(axes[0, 0], hospital_exposed_edges, "A", base_network, basemap_alpha=0.4, bounds_3857=bounds_3857)
    _plot_impact_panel(axes[0, 1], factory_exposed_edges, "B", base_network, basemap_alpha=0.4, bounds_3857=bounds_3857)
    _plot_impact_panel(axes[1, 0], police_exposed_edges, "C", base_network, basemap_alpha=0.4, bounds_3857=bounds_3857)
    _plot_impact_panel(axes[1, 1], fire_exposed_edges, "D", base_network, basemap_alpha=0.4, bounds_3857=bounds_3857)

    legend_handles = [Patch(facecolor=IMPACT_COLORS[i], label=f"{IMPACT_LABELS[i]} delay",
                            edgecolor="darkred", linewidth=0.5)
                      for i in range(len(IMPACT_LABELS))]
    fig.legend(handles=legend_handles, title="Increased travel time", loc="lower center",
               bbox_to_anchor=(0.5, 0.0), ncol=len(IMPACT_LABELS), fontsize=10,
               title_fontsize=12, frameon=True, fancybox=True, framealpha=0.9)

    plt.savefig(Path(figure_path) / "criticality_2x2.png", dpi=dpi, bbox_inches="tight")
    _show_or_close(show_figures)


def save_impact_layers(
    impact_layers: dict[str, gpd.GeoDataFrame | None],
    parquet_dir: Path,
    gdb_path: Path,
    output_crs: str = "EPSG:6316",
) -> None:
    """Save each accessibility impact layer to parquet + GDB layer + Excel."""
    for layer_name, gdf in impact_layers.items():
        if gdf is None or len(gdf) == 0:
            print(f"Skipping {layer_name}: no data")
            continue
        save_criticality_vector(gdf, parquet_dir, gdb_path, layer_name, output_crs)


# ===========================================================================
# 5d — Combined climate criticality
# ===========================================================================
#
# The 5d workflow assembles every hazard, national-disruption and local-
# accessibility result into a single per-road climate-criticality index.
# Pipeline: load_and_preprocess -> clean_data -> prepare_metrics ->
# deduplicate_by_section -> score_climate_criticality -> plots / stats / export.
#
# Two behaviours are selectable (wired to NetworkConfig in the script, hardcoded
# in the notebook):
#   * climate_hazards_only — build the hazard sub-index (H) only from the
#     climate-change-driven hazards, or from all mapped hazards.
#   * normalize_subindices — combine min-max normalised sub-indices (CC_norm)
#     or the raw summed convex scores (CC_raw).


def add_impact_column(
    base_gdf: gpd.GeoDataFrame,
    edges_gdf: gpd.GeoDataFrame,
    impact_col_name: Any,
    predicate: str = "intersects",
    agg: str = "mean",
    col_to_focus: str = "travel_time_impact",
) -> gpd.GeoDataFrame:
    """Spatially join *edges_gdf* to *base_gdf*, aggregate *col_to_focus* per base
    index, and attach it as ``impact_col_name``.

    - ``predicate``: 'intersects', 'within', 'contains', 'touches'.
    - ``agg``: 'mean', 'max', 'min', 'median', etc.
    """
    edges = edges_gdf[[col_to_focus, "geometry"]].copy()
    joined = base_gdf.sjoin(edges, how="left", predicate=predicate)
    agg_series = joined.groupby(joined.index)[col_to_focus].agg(agg)
    base_gdf[impact_col_name] = agg_series.reindex(base_gdf.index)
    return base_gdf


def load_and_preprocess_criticality_data(
    *,
    hazard_exposure_path: Path,
    criticality_results_path: Path,
    hospital_impacts_path: Path,
    factory_impacts_path: Path,
    police_impacts_path: Path,
    fire_impacts_path: Path,
    border_impacts_path: Path,
    port_impacts_path: Path,
    railway_impacts_path: Path,
    future_floods_change_rp_path: Path,
    future_rainfall_change_path: Path,
) -> gpd.GeoDataFrame:
    """Assemble the nationwide road network enriched with every per-topic metric.

    Reads the hazard-exposure layer (4a/5a) and the full criticality results (2),
    joins them so no critical road is dropped, then attaches accessibility delays
    (hospital/fire/police/factory/port/border/railway), the projected change in
    flood return period, and the projected change in extreme rainfall via
    ``add_impact_column(..., predicate='intersects', agg='mean')``. ``max_pavement_temp``
    and ``wildfire_susc`` are already present in the hazard-exposure layer and are
    carried through unchanged. All inputs are aligned to the hazard layer CRS.
    """
    gdf_hazards = gpd.read_parquet(hazard_exposure_path)

    gdf_all_critical = gpd.read_parquet(criticality_results_path).to_crs(gdf_hazards.crs)

    for col in ["oznaka_deo", "oznaka_put", "pocetna_st", "zavrsna_st"]:
        print(f"\n--- {col} ---")
        print(f"  gdf_all_critical: {gdf_all_critical[col].nunique()} unique / {len(gdf_all_critical)} total, nulls: {gdf_all_critical[col].isna().sum()}")
        print(f"  gdf_hazards:      {gdf_hazards[col].nunique()} unique / {len(gdf_hazards)} total, nulls: {gdf_hazards[col].isna().sum()}")

    # Merge so that roads only in the criticality results are also included.
    # A left join on gdf_all_critical ensures no critical road is dropped, while
    # hazard columns are NaN for roads with no hazard exposure. Keep only the
    # hazard-specific columns from gdf_hazards to avoid conflicts.
    hazard_only_cols = [
        col for col in gdf_hazards.columns
        if col not in gdf_all_critical.columns or col == "geometry"
    ]

    crs = gdf_all_critical.crs  # save before the join

    gdf_all_critical = gdf_all_critical.drop(columns=["index_right"], errors="ignore")
    gdf_hazards = gdf_hazards.drop(columns=["index_right"], errors="ignore")

    gdf_hazards_joined = gpd.sjoin(
        gdf_all_critical,
        gdf_hazards[hazard_only_cols].assign(_hazard_len=gdf_hazards.geometry.length),
        how="left",
        predicate="intersects",
    ).drop(columns=["index_right"], errors="ignore")

    gdf_hazards = (
        gdf_hazards_joined
        .sort_values("_hazard_len", ascending=False)
        .groupby(level=0)
        .first()
        .drop(columns="_hazard_len")
    )

    gdf_hazards = gpd.GeoDataFrame(gdf_hazards, geometry="geometry", crs=crs)
    assert len(gdf_hazards) == len(gdf_all_critical)

    # Accessibility impact layers (5c) and climate-change layers (4b)
    hospital_exposed_edges = gpd.read_parquet(hospital_impacts_path).to_crs(gdf_hazards.crs)
    factory_exposed_edges = gpd.read_parquet(factory_impacts_path).to_crs(gdf_hazards.crs)
    police_exposed_edges = gpd.read_parquet(police_impacts_path).to_crs(gdf_hazards.crs)
    fire_exposed_edges = gpd.read_parquet(fire_impacts_path).to_crs(gdf_hazards.crs)
    border_exposed_edges = gpd.read_parquet(border_impacts_path).to_crs(gdf_hazards.crs)
    port_exposed_edges = gpd.read_parquet(port_impacts_path).to_crs(gdf_hazards.crs)
    railway_exposed_edges = gpd.read_parquet(railway_impacts_path).to_crs(gdf_hazards.crs)

    future_flood_change_rp = gpd.read_parquet(future_floods_change_rp_path).to_crs(gdf_hazards.crs)
    future_rainfall_change = gpd.read_parquet(future_rainfall_change_path).to_crs(gdf_hazards.crs)

    gdf_hazards = add_impact_column(gdf_hazards, hospital_exposed_edges, "hospital_delay")
    gdf_hazards = add_impact_column(gdf_hazards, factory_exposed_edges, "factory_delay")
    gdf_hazards = add_impact_column(gdf_hazards, police_exposed_edges, "police_delay")
    gdf_hazards = add_impact_column(gdf_hazards, fire_exposed_edges, "fire_delay")
    gdf_hazards = add_impact_column(gdf_hazards, port_exposed_edges, "port_delay")
    gdf_hazards = add_impact_column(gdf_hazards, border_exposed_edges, "border_delay")
    gdf_hazards = add_impact_column(gdf_hazards, railway_exposed_edges, "railway_delay")
    gdf_hazards = add_impact_column(
        gdf_hazards, future_flood_change_rp, "future_flood_change", col_to_focus="rp30_mean"
    )
    gdf_hazards = add_impact_column(
        gdf_hazards, future_rainfall_change, "future_rainfall_change", col_to_focus="max_rx1day_pct"
    )

    return gdf_hazards


def clean_data(gdf_hazards: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Remove invalid, duplicate, and negligibly short road-hazard segments.

    Steps: drop missing ``oznaka_deo``; drop exact duplicates (incl. geometry);
    drop duplicates on non-geometry columns; drop segments < 0.05 km whose
    ``oznaka_deo`` appears more than once, provided the group keeps a survivor
    so no unique ``oznaka_deo`` is lost entirely.
    """
    print(f"Rows before cleaning: {len(gdf_hazards)}")

    gdf_hazards = gdf_hazards.dropna(subset=["oznaka_deo"])
    print(f"Rows after dropping invalid oznaka_deo: {len(gdf_hazards)}")

    gdf_hazards = gdf_hazards.drop_duplicates()
    print(f"Rows after dropping exact duplicates (including geometry): {len(gdf_hazards)}")

    non_geom_cols = [col for col in gdf_hazards.columns if col != "geometry"]
    gdf_hazards = gdf_hazards.drop_duplicates(subset=non_geom_cols)
    print(f"Rows after dropping duplicates (excluding geometry): {len(gdf_hazards)}")

    duplicated_mask = gdf_hazards["oznaka_deo"].duplicated(keep=False)
    short_mask = gdf_hazards["road_length"] < 0.05
    has_survivor = gdf_hazards.groupby("oznaka_deo")["road_length"].transform(
        lambda x: (x >= 0.05).any()
    )
    drop_mask = duplicated_mask & short_mask & has_survivor

    gdf_hazards = gdf_hazards[~drop_mask]
    print(f"Rows after dropping sections with double oznaka_deo shorter than 50m: {len(gdf_hazards)}")

    return gdf_hazards


def prepare_metrics(gdf_hazards: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Rename source columns to standard metric names and derive hazard indicators.

    Renames ``max_depth`` -> ``flood_depth``, ``dužina_sn`` -> ``snow_drift``,
    ``datum_evid`` -> ``landslide_date``, and ``future_flood_change`` (the raw
    future return period from ``rp30_mean``) -> ``future_flood_return_period``.
    Derives ``future_flood_change`` = 100 - return period (a shorter future RP
    means higher flood risk; the raw column is preserved for auditability and
    is used as the MIN aggregator during deduplication). Builds binary
    ``landslide_exposure`` and clips ``wildfire_susc`` to a 0/1
    ``wildfire_susceptibility`` indicator.
    """
    gdf_hazards = gdf_hazards.copy()
    gdf_hazards.index.name = "section_id"

    gdf_hazards = gdf_hazards.rename(
        columns={
            "max_depth": "flood_depth",
            "dužina_sn": "snow_drift",
            "datum_evid": "landslide_date",
            "future_flood_change": "future_flood_return_period",
        }
    )

    gdf_hazards["future_flood_change"] = 100 - gdf_hazards["future_flood_return_period"]

    gdf_hazards["landslide_exposure"] = np.where(
        gdf_hazards["landslide_date"].notna()
        & (gdf_hazards["landslide_date"].astype(str).str.strip() != ""),
        1.0,
        0.0,
    )

    gdf_hazards["wildfire_susceptibility"] = gdf_hazards["wildfire_susc"].clip(0, 1).astype(float)

    return gdf_hazards


# ---------------------------------------------------------------------------
# 5d — Deduplication of road sections (one row per oznaka_deo)
# ---------------------------------------------------------------------------

# Attributes taken from the row with the maximum road_length within each group
# (road_length itself is summed). Geometry is taken from that same longest row.
_DEDUP_TAKE_FROM_LONGEST = [
    "section_id", "from_id", "to_id", "objectid", "oznaka_deo", "smer_gdf1",
    "kategorija", "oznaka_put", "oznaka_poc", "naziv_poce", "oznaka_zav",
    "naziv_zavr", "duzina_deo", "pocetna_st", "zavrsna_st", "stanje", "id",
    "passenger_cars", "buses", "light_trucks", "medium_trucks", "heavy_trucks",
    "articulated_vehicles", "total_aadt", "road_length", "speed", "fft",
]
_CRITICALITY_ORDER = ["No criticality", "Very Low", "Low", "Medium", "High", "Very High"]


def deduplicate_by_section(
    gdf: gpd.GeoDataFrame,
    group_col: str = "oznaka_deo",
    length_col: str = "road_length",
    min_col: str = "future_flood_return_period",
    sort_col: str = "criticality_mean",
    weighted_avg_cols: list[str] | None = None,
) -> gpd.GeoDataFrame:
    """Collapse multiple segments sharing an ``oznaka_deo`` into one representative row.

    Aggregation rules (ported from ``road_processing.ipynb``):
      * identity / network / traffic attributes — taken from the row with the
        maximum ``road_length`` in the group (and so is the geometry);
      * ``road_length`` — summed;
      * ``min_col`` (future flood return period) — minimum (worst future risk);
      * class columns (``H_class``/``T_class``/``A_class``/``CC_class``), if
        present — the highest criticality label;
      * any ``weighted_avg_cols`` — road-length-weighted mean;
      * every remaining metric column — maximum (worst case).
    ``*_norm`` / ``*_log`` / ``mean_class`` helper columns are dropped first.
    Run BEFORE scoring so normalisation reflects the deduplicated distribution.
    """
    gdf = gdf.copy()
    geom_name = gdf.geometry.name

    drop_cols = [
        c for c in gdf.columns
        if c.endswith("_norm") or c.endswith("_log") or c == "mean_class"
    ]
    gdf = gdf.drop(columns=drop_cols, errors="ignore")

    weighted_avg_cols = [c for c in (weighted_avg_cols or []) if c in gdf.columns]
    take_longest = [c for c in _DEDUP_TAKE_FROM_LONGEST if c in gdf.columns]
    class_cols = [c for c in ("H_class", "T_class", "A_class", "CC_class") if c in gdf.columns]
    special = set(take_longest) | {min_col, group_col, geom_name} | set(weighted_avg_cols) | set(class_cols)
    max_cols = [c for c in gdf.columns if c not in special]

    def aggregate_group(grp):
        if len(grp) == 1:
            return grp

        result = {group_col: grp[group_col].iloc[0]}
        base_row = grp.loc[grp[length_col].idxmax()]
        result[geom_name] = base_row[geom_name]

        for col in take_longest:
            if col == length_col:
                result[col] = grp[col].sum()
            elif col != group_col:
                result[col] = base_row[col]

        if min_col in grp.columns:
            result[min_col] = grp[min_col].min()

        for col in class_cols:
            non_null = grp[col].dropna()
            result[col] = (
                pd.Categorical(non_null, categories=_CRITICALITY_ORDER, ordered=True).max()
                if not non_null.empty else np.nan
            )

        for col in max_cols:
            try:
                result[col] = grp[col].max()
            except TypeError:
                nn = grp[col].dropna()
                result[col] = nn.iloc[0] if not nn.empty else np.nan

        weights = grp[length_col].fillna(0)
        for col in weighted_avg_cols:
            values = grp[col]
            mask = values.notna() & weights.notna() & (weights > 0)
            if mask.any():
                result[col] = (values[mask] * weights[mask]).sum() / weights[mask].sum()
            else:
                result[col] = values.mean() if weights.sum() == 0 else np.nan

        return pd.DataFrame([result])

    n_before = len(gdf)
    merged = (
        gdf.groupby(group_col, sort=False, group_keys=False)
        .apply(aggregate_group)
        .reset_index(drop=True)
    )
    merged = merged[gdf.columns]
    merged = gpd.GeoDataFrame(merged, geometry=geom_name, crs=gdf.crs)

    if sort_col in merged.columns:
        merged = merged.sort_values(sort_col, ascending=False).reset_index(drop=True)

    print(
        f"Deduplicated by {group_col}: {n_before} -> {len(merged)} rows "
        f"({n_before - len(merged)} collapsed; {merged[group_col].duplicated().sum()} remaining dups)"
    )
    return merged


# ---------------------------------------------------------------------------
# 5d — Climate-criticality scoring (convex quintile + multiplicative index)
# ---------------------------------------------------------------------------

# Hazard Exposure (H). The climate subset is used when climate_hazards_only=True.
_HAZARD_METRICS = [
    "flood_depth", "future_rainfall_change", "future_flood_change",
    "max_pavement_temp", "wildfire_susceptibility", "landslide_exposure", "snow_drift",
]
_HAZARD_CLIMATE_METRICS = ["future_rainfall_change", "max_pavement_temp", "landslide_exposure"]
# National-scale travel disruption (T)
_TRAVEL_METRICS = ["phl", "thl", "pkl", "tkl"]
# Local accessibility (A)
_ACCESSIBILITY_METRICS = [
    "hospital_delay", "fire_delay", "police_delay", "factory_delay",
    "port_delay", "border_delay", "railway_delay",
]
# Continuous metrics that get a log(x+1) transform before normalisation.
# Binary indicators (landslide_exposure, wildfire_susceptibility) are excluded.
_LOG_TRANSFORM_METRICS = [
    "flood_depth", "future_rainfall_change", "future_flood_change",
    "snow_drift", "max_pavement_temp",
    "phl", "thl", "pkl", "tkl",
    "hospital_delay", "fire_delay", "police_delay",
    "factory_delay", "port_delay", "border_delay", "railway_delay",
]
# Convex score per quintile: Q1=0, Q2=1, Q3=2, Q4=5, Q5=10 (0 = no criticality).
CONVEX_MAP = {0: 0, 1: 0, 2: 1, 3: 2, 4: 5, 5: 10}


def _safe_minmax(series: pd.Series) -> pd.Series:
    """Min-max normalise to [0, 1]; zeros if constant or empty."""
    s = series.astype(float).fillna(0.0)
    mn, mx = s.min(), s.max()
    if pd.isna(mn) or pd.isna(mx) or mx == mn:
        return pd.Series(np.zeros(len(s)), index=s.index)
    return (s - mn) / (mx - mn)


def _norm01(series: pd.Series) -> pd.Series:
    """Min-max normalise a sub-index to [0, 1] (zeros if constant)."""
    mn, mx = series.min(), series.max()
    if mx == mn:
        return pd.Series(0.0, index=series.index)
    return (series - mn) / (mx - mn)


def _quintile_score(series: pd.Series) -> pd.Series:
    """Percentile-rank non-zero values into integer quintiles 1-5 (0 stays 0)."""
    result = pd.Series(0, index=series.index, dtype=int)
    non_zeros = series[series != 0]
    if not non_zeros.empty:
        ranks = non_zeros.rank(pct=True)
        result[series != 0] = np.ceil(ranks * 5).clip(1, 5).astype(int)
    return result


def _classify(series: pd.Series) -> pd.Series:
    """Label non-zero values by quintile; zeros become 'No criticality'."""
    labels_map = {1: "Very Low", 2: "Low", 3: "Medium", 4: "High", 5: "Very High"}
    result = pd.Series("No criticality", index=series.index, dtype="object")
    non_zeros = series[series != 0]
    if not non_zeros.empty:
        ranks = non_zeros.rank(pct=True)
        bins = np.ceil(ranks * 5).clip(1, 5).astype(int)
        result[series != 0] = bins.map(labels_map)
    return result


def score_climate_criticality(
    gdf: gpd.GeoDataFrame,
    climate_hazards_only: bool = True,
    normalize_subindices: bool = True,
) -> gpd.GeoDataFrame:
    """Compute the combined climate-criticality index per road section.

    Each metric is (optionally log-transformed,) min-max normalised, ranked into
    quintiles and mapped to a convex score (``CONVEX_MAP``). Convex scores are
    summed within each sub-index — Hazard Exposure (H), Travel Disruption (T),
    Local Accessibility (A) — then combined multiplicatively:

      * ``normalize_subindices=True``  -> CC = norm(H) x (norm(T) + norm(A))   [CC_norm]
      * ``normalize_subindices=False`` -> CC = H x (T + A)                     [CC_raw]

    ``climate_hazards_only`` selects the climate-only hazard subset
    (rainfall change, pavement temperature, landslide) for H instead of all
    hazards. Both ``H_all`` and ``H_climate`` are always computed and kept for
    transparency; ``H`` is the selected one. The chosen index is stored as
    ``climate_criticality`` (+ ``climate_criticality_class``), with
    ``criticality_mean`` / ``mean_class`` kept as aliases for the plotting and
    statistics helpers.
    """
    gdf = gdf.copy()
    all_metrics = _HAZARD_METRICS + _TRAVEL_METRICS + _ACCESSIBILITY_METRICS

    for col in all_metrics:
        if col not in gdf.columns:
            print(f"Warning: metric '{col}' not found — creating with zeros.")
            gdf[col] = 0.0
        gdf[col] = pd.to_numeric(gdf[col], errors="coerce").fillna(0.0)

    # Working columns: log(x+1) where specified (log1p(0)=0 preserves zeros).
    working = {}
    for col in all_metrics:
        if col in _LOG_TRANSFORM_METRICS:
            wcol = f"{col}_log"
            gdf[wcol] = np.log1p(gdf[col])
            working[col] = wcol
        else:
            working[col] = col

    # Negative future changes (drier / longer RP) are not a risk driver.
    for col in ["future_rainfall_change", "future_flood_change"]:
        gdf[working[col]] = gdf[working[col]].clip(lower=0)

    # Normalise -> quintile -> convex score per metric.
    for col in all_metrics:
        gdf[f"{col}_norm"] = _safe_minmax(gdf[working[col]])
        gdf[f"{col}_q"] = _quintile_score(gdf[f"{col}_norm"])
        gdf[f"{col}_cv"] = gdf[f"{col}_q"].map(CONVEX_MAP)

    # Sub-indices = sum of convex scores.
    gdf["H_all"] = gdf[[f"{c}_cv" for c in _HAZARD_METRICS]].sum(axis=1)
    gdf["H_climate"] = gdf[[f"{c}_cv" for c in _HAZARD_CLIMATE_METRICS]].sum(axis=1)
    gdf["T"] = gdf[[f"{c}_cv" for c in _TRAVEL_METRICS]].sum(axis=1)
    gdf["A"] = gdf[[f"{c}_cv" for c in _ACCESSIBILITY_METRICS]].sum(axis=1)

    hazard_col = "H_climate" if climate_hazards_only else "H_all"
    gdf["H"] = gdf[hazard_col]

    def _combine(hazard_index: pd.Series) -> pd.Series:
        """Combine a hazard sub-index with T and A (multiplicative)."""
        if normalize_subindices:
            return _norm01(hazard_index) * (_norm01(gdf["T"]) + _norm01(gdf["A"]))
        return hazard_index * (gdf["T"] + gdf["A"])

    # Primary (selected) track — drives the maps, statistics and the
    # non-extended Excel sheets.
    gdf["climate_criticality"] = _combine(gdf["H"])
    gdf["H_class"] = _classify(gdf["H"])
    gdf["T_class"] = _classify(gdf["T"])
    gdf["A_class"] = _classify(gdf["A"])
    gdf["climate_criticality_class"] = _classify(gdf["climate_criticality"])

    # Extended track — always uses ALL hazards (H_all). Feeds the
    # "Overview Extended" / "Hazard Exposure Extended" Excel sheets.
    gdf["H_extended"] = gdf["H_all"]
    gdf["H_class_extended"] = _classify(gdf["H_all"])
    gdf["climate_criticality_extended"] = _combine(gdf["H_all"])
    gdf["climate_criticality_class_extended"] = _classify(gdf["climate_criticality_extended"])

    # Aliases consumed by the plotting / statistics helpers below.
    gdf["criticality_mean"] = gdf["climate_criticality"]
    gdf["mean_class"] = gdf["climate_criticality_class"]
    gdf["CC_class"] = gdf["climate_criticality_class"]

    variant = "CC_norm" if normalize_subindices else "CC_raw"
    print(
        f"Scored climate criticality (hazard='{hazard_col}', {variant}): "
        f"mean={gdf['climate_criticality'].mean():.4f}, max={gdf['climate_criticality'].max():.4f}"
    )
    return gdf


# ---------------------------------------------------------------------------
# 5d — Maps
# ---------------------------------------------------------------------------

_CC_LABELS = ["No criticality", "Very Low", "Low", "Medium", "High", "Very High"]


def plot_climate_criticality_components(
    gdf_hazards: gpd.GeoDataFrame,
    figure_path: Path,
    gdb_path: Path,
    lyrx_dir: Path,
    show_figures: bool = True,
) -> None:
    """Three-panel map of the H / T / A sub-index classes, plus ArcGIS layers.

    Each sub-index is written as a feature class in *gdb_path* (the lyrx data
    source) with a matching .lyrx in *lyrx_dir*; no GeoPackage is produced.
    """
    class_cols = ["H_class", "T_class", "A_class"]
    labels = _CC_LABELS
    colors = ["#e0e0e0", "#edf8fb", "#b3cde3", "#8c96c6", "#8856a7", "#810f7c"]
    color_map = dict(zip(labels, colors))
    width_mapping = {
        "No criticality": 0.2, "Very Low": 0.6, "Low": 0.9,
        "Medium": 1.3, "High": 2.0, "Very High": 3.0,
    }

    original_crs = gdf_hazards.crs
    gdf_hazards = gdf_hazards.to_crs(epsg=3857)

    def plot_panel(ax, gdf, class_col, letter, title):
        for cat in labels:
            subset = gdf[gdf[class_col] == cat]
            if not subset.empty:
                subset.plot(
                    ax=ax, color=color_map[cat], linewidth=width_mapping[cat],
                    alpha=0.8 if cat != "No criticality" else 0.4,
                    zorder=labels.index(cat),
                )
        cx.add_basemap(ax, source=cx.providers.CartoDB.Positron, attribution=False)
        ax.axis("off")
        ax.text(0.05, 0.95, letter, transform=ax.transAxes, fontsize=22,
                fontweight="bold", verticalalignment="top",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.9))
        ax.set_title(title, fontsize=14, fontweight="bold")

    fig, axes = plt.subplots(1, 3, figsize=(15, 7))
    titles = ["Hazard-Exposure", "National-Scale Travel Disruption", "Local Accessibility"]
    for i, ax in enumerate(axes):
        plot_panel(axes[i], gdf_hazards, class_cols[i], chr(65 + i), titles[i])

    legend_handles = [
        Patch(facecolor=color_map[lbl], label=lbl, edgecolor="grey", linewidth=0.5)
        for lbl in labels
    ]
    fig.legend(handles=legend_handles, title="Criticality Level", loc="lower center",
               bbox_to_anchor=(0.5, 0.02), ncol=6, fontsize=13, title_fontsize=15, frameon=True)

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.18)
    plt.savefig(Path(figure_path) / "criticality_analysis_3panel.png", dpi=300, bbox_inches="tight")

    # ArcGIS layers per sub-index (restore original CRS first)
    gdf_hazards = gdf_hazards.to_crs(original_crs)
    for field, name, title in [
        ("H_class", "hazard_exposure", "Hazard Exposure"),
        ("T_class", "travel_disruption", "National-Scale Travel Disruption"),
        ("A_class", "local_accessibility", "Local Accessibility"),
    ]:
        save_lyrx_layer(
            gdf=gdf_hazards, gpkg_path=None, gdb_path=gdb_path,
            lyrx_path=Path(lyrx_dir) / f"{name}.lyrx", layer_name=name,
            labels=labels, colors=colors, width_mapping=width_mapping,
            field=field, title=title,
        )

    _show_or_close(show_figures)


def plot_combined_climate_criticality(
    gdf_hazards: gpd.GeoDataFrame,
    figure_path: Path,
    gdb_path: Path,
    lyrx_dir: Path,
    cc_class_col: str = "climate_criticality_class",
    show_figures: bool = True,
) -> None:
    """Single map of the combined climate-criticality class, plus an ArcGIS layer.

    The layer is written as a feature class in *gdb_path* (the lyrx data source)
    with a matching .lyrx in *lyrx_dir*; no GeoPackage is produced.
    """
    labels = _CC_LABELS
    colors = ["#e0e0e0", "#ffffcc", "#a1dab4", "#41b6c4", "#2c7fb8", "#253494"]
    color_map = dict(zip(labels, colors))
    width_mapping = {
        "No criticality": 0.2, "Very Low": 0.6, "Low": 1.0,
        "Medium": 1.5, "High": 2.2, "Very High": 3.2,
    }

    original_crs = gdf_hazards.crs
    gdf_hazards = gdf_hazards.to_crs(epsg=3857)

    fig, ax = plt.subplots(1, 1, figsize=(7, 9))
    for cat in labels:
        subset = gdf_hazards[gdf_hazards[cc_class_col] == cat]
        if not subset.empty:
            subset.plot(
                ax=ax, color=color_map[cat], linewidth=width_mapping[cat],
                alpha=0.8 if cat != "No criticality" else 0.4, zorder=labels.index(cat),
            )

    cx.add_basemap(ax, source=cx.providers.CartoDB.Positron, attribution=False)
    ax.axis("off")
    legend_handles = [
        Patch(facecolor=color_map[lbl], label=lbl, edgecolor="grey", linewidth=0.5)
        for lbl in labels
    ]
    ax.legend(handles=legend_handles, title="Climate Criticality", loc="upper right",
              fontsize=14, title_fontsize=16, frameon=True, framealpha=0.9)

    plt.tight_layout()
    plt.savefig(Path(figure_path) / "climate_criticality_mean.png", dpi=150, bbox_inches="tight")

    gdf_hazards = gdf_hazards.to_crs(original_crs)
    save_lyrx_layer(
        gdf=gdf_hazards, gpkg_path=None, gdb_path=gdb_path,
        lyrx_path=Path(lyrx_dir) / "climate_criticality_index.lyrx",
        layer_name="climate_criticality", labels=labels, colors=colors,
        width_mapping=width_mapping, field=cc_class_col, title="Climate Criticality Metric",
    )

    _show_or_close(show_figures)


# ---------------------------------------------------------------------------
# 5d — Summary statistics
# ---------------------------------------------------------------------------

def print_climate_criticality_statistics(gdf_hazards: gpd.GeoDataFrame) -> None:
    """Print distribution, road-category, tier, top-section and correlation summaries."""
    gdf_hazards = gdf_hazards.copy()
    class_cols = ["H_class", "T_class", "A_class", "climate_criticality_class"]
    class_names = ["Hazard Exposure", "Travel Disruption", "Local Accessibility", "Combined Criticality"]
    index_cols = ["H", "T", "A", "climate_criticality"]
    labels = _CC_LABELS
    high_crit_labels = ["High", "Very High"]

    print("=" * 80)
    print("CLIMATE CRITICALITY SUMMARY STATISTICS")
    print("=" * 80)

    print("\n1. OVERALL DISTRIBUTION BY CRITICALITY CLASS")
    for class_col, class_name in zip(class_cols, class_names):
        print(f"\n{class_name}:")
        counts = gdf_hazards[class_col].value_counts()
        total = len(gdf_hazards)
        for label in labels:
            count = int(counts.get(label, 0))
            print(f"  {label:15s}: {count:5d} ({count / total * 100:5.1f}%)")

    print("\n2. MEAN CRITICALITY SCORES BY ROAD CATEGORY")
    road_summary = gdf_hazards.groupby("kategorija").agg(
        {
            "H": ["count", "mean", "median", "max"],
            "T": ["mean", "median", "max"],
            "A": ["mean", "median", "max"],
            "climate_criticality": ["mean", "median", "max"],
        }
    ).round(4)
    print(road_summary.to_string())

    print("\n3. HIGH-TIER vs LOW-TIER ROADS COMPARISON")
    high_tier = ["IA", "IM", "IB"]
    low_tier = ["IIA", "IIB"]
    gdf_hazards["road_tier"] = np.where(
        gdf_hazards["kategorija"].isin(high_tier), "High-Tier (IA/IM/IB)",
        np.where(gdf_hazards["kategorija"].isin(low_tier), "Low-Tier (IIA/IIB)", "Other"),
    )
    tier_summary = gdf_hazards.groupby("road_tier")[index_cols].mean().round(4)
    print(tier_summary.to_string())

    high_tier_scores = gdf_hazards.loc[gdf_hazards["road_tier"] == "High-Tier (IA/IM/IB)", "climate_criticality"]
    low_tier_scores = gdf_hazards.loc[gdf_hazards["road_tier"] == "Low-Tier (IIA/IIB)", "climate_criticality"]
    if len(high_tier_scores) > 0 and len(low_tier_scores) > 0:
        stat, pvalue = scipy_stats.mannwhitneyu(high_tier_scores, low_tier_scores, alternative="two-sided")
        print("\nMann-Whitney U test (High-Tier vs Low-Tier):")
        print(f"  U-statistic: {stat:.2f}  p-value: {pvalue:.6f}  Significant: {'Yes' if pvalue < 0.05 else 'No'}")

    print("\n4. TOP 10 MOST CRITICAL SECTIONS (Combined)")
    top_cols = [c for c in [
        "oznaka_deo", "kategorija", "oznaka_put", "naziv_poce", "naziv_zavr",
        "H", "T", "A", "climate_criticality", "climate_criticality_class",
    ] if c in gdf_hazards.columns]
    print(gdf_hazards.nlargest(10, "climate_criticality")[top_cols].to_string())

    print("\n5. CORRELATION BETWEEN SUB-INDICES")
    print(gdf_hazards[index_cols].corr().round(3).to_string())

    print("\n6. SHARE OF HIGH/VERY HIGH CRITICALITY BY ROAD CATEGORY")
    for class_col, class_name in zip(class_cols, class_names):
        print(f"\n{class_name}:")
        for cat in sorted(gdf_hazards["kategorija"].dropna().unique()):
            subset = gdf_hazards[gdf_hazards["kategorija"] == cat]
            count_high = subset[class_col].isin(high_crit_labels).sum()
            total = len(subset)
            print(f"  {cat:12s} | {count_high / total * 100:6.1f}% | {count_high:5d} / {total}")

    total_sections = len(gdf_hazards)
    high_vh = gdf_hazards["climate_criticality_class"].isin(high_crit_labels).sum()
    print(f"\nKEY: {total_sections} sections analysed; "
          f"{high_vh} ({high_vh / total_sections * 100:.1f}%) High/Very High combined criticality.")


# ---------------------------------------------------------------------------
# 5d — Excel and geospatial export
# ---------------------------------------------------------------------------

_EXCEL_ID_COLS = [
    "objectid", "oznaka_deo", "smer_gdf1", "kategorija", "oznaka_put",
    "oznaka_poc", "naziv_poce", "oznaka_zav", "naziv_zavr",
    "duzina_deo", "pocetna_st", "zavrsna_st", "stanje",
]
_EXCEL_TRAFFIC_COLS = [
    "passenger_cars", "buses", "light_trucks", "medium_trucks",
    "heavy_trucks", "articulated_vehicles", "total_aadt",
]


def _format_climate_workbook(path: Path) -> None:
    """Apply header styling, class colour-coding, widths and freeze panes."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(bottom=Side(style="thin", color="D5D8DC"))
    class_fills = {
        "No criticality": PatternFill("solid", fgColor="E0E0E0"),
        "Very Low": PatternFill("solid", fgColor="EDF8FB"),
        "Low": PatternFill("solid", fgColor="B3CDE3"),
        "Medium": PatternFill("solid", fgColor="8C96C6"),
        "High": PatternFill("solid", fgColor="8856A7"),
        "Very High": PatternFill("solid", fgColor="810F7C"),
    }
    class_fonts_white = {"High", "Very High"}
    class_columns = {
        "H_class", "H_climate_class", "T_class", "A_class",
        "CC_class", "climate_criticality_class",
        "H_class_extended", "climate_criticality_class_extended",
    }

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        headers = {cell.value: cell.column for cell in ws[1]}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
        for col_name in class_columns:
            if col_name in headers:
                col_idx = headers[col_name]
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = str(cell.value) if cell.value else ""
                    if val in class_fills:
                        cell.fill = class_fills[val]
                    if val in class_fonts_white:
                        cell.font = Font(name="Arial", size=10, color="FFFFFF")
        # The Metric Descriptions sheet holds long text, so allow wider columns.
        width_cap = 90 if ws.title == "Metric Descriptions" else 30
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, width_cap)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# Static descriptions for the human-readable "Metric Descriptions" sheet.
_DESC_IDENTIFIERS = [
    ("objectid", "Sequential feature identifier", "—"),
    ("oznaka_deo", "Road section identifier (deonica)", "—"),
    ("smer_gdf1", "Carriageway / direction indicator", "—"),
    ("kategorija", "Road category (IA, IB, IM, IIA, IIB)", "—"),
    ("oznaka_put", "Road (route) designation", "—"),
    ("oznaka_poc", "Start node code", "—"),
    ("naziv_poce", "Start node name", "—"),
    ("oznaka_zav", "End node code", "—"),
    ("naziv_zavr", "End node name", "—"),
    ("duzina_deo", "Official section length (source dataset)", "km"),
    ("pocetna_st", "Start stationing (chainage)", "km"),
    ("zavrsna_st", "End stationing (chainage)", "km"),
    ("stanje", "Section condition / status", "—"),
]
_DESC_TRAFFIC = [
    ("passenger_cars", "Daily passenger car count", "vehicles/day"),
    ("buses", "Daily bus count", "vehicles/day"),
    ("light_trucks", "Daily light truck count", "vehicles/day"),
    ("medium_trucks", "Daily medium truck count", "vehicles/day"),
    ("heavy_trucks", "Daily heavy truck count", "vehicles/day"),
    ("articulated_vehicles", "Daily articulated vehicle count", "vehicles/day"),
    ("total_aadt", "Total annual average daily traffic", "vehicles/day"),
]
_DESC_HAZARD = [
    ("flood_depth", "Maximum inundation depth along section (1-in-100-year flood)", "cm"),
    ("future_rainfall_change", "Projected change in extreme rainfall intensity (RCP 8.5, far future)", "% change"),
    ("future_flood_change", "Projected change in river flood return period under 4°C of global warming (100 baseline return period)", "years of return period"),
    ("max_pavement_temp", "Projected maximum pavement temperature (hottest 7 days)", "°C"),
    ("wildfire_susceptibility", "Section intersects high-risk wildfire zone", "0/1 (no/yes)"),
    ("landslide_exposure", "Section intersects historical landslide zone", "0/1 (no/yes)"),
    ("snow_drift", "Length of historical snow drift intersecting this section", "km"),
]
_DESC_TRAVEL = [
    ("phl", "Passenger hours lost (daily, weighted by traffic volume)", "passenger-hours"),
    ("thl", "Tonnage hours lost (daily, weighted by freight volume)", "ton-hours"),
    ("pkl", "Passenger kilometers lost (additional distance due to rerouting)", "passenger-km"),
    ("tkl", "Tonnage kilometers lost (additional freight distance)", "ton-km"),
]
_DESC_ACCESS = [
    ("hospital_delay", "Additional travel time to nearest hospital", "hours"),
    ("fire_delay", "Additional travel time from fire stations to service areas", "hours"),
    ("police_delay", "Additional travel time from police stations to service areas", "hours"),
    ("factory_delay", "Additional travel time from industrial areas to border crossings", "hours"),
    ("port_delay", "Additional travel time from agricultural areas to ports", "hours"),
    ("border_delay", "Additional travel time from agricultural areas to border crossings", "hours"),
    ("railway_delay", "Additional travel time from agricultural areas to railway stations", "hours"),
]


def _metric_descriptions_df(
    climate_hazards_only: bool, normalize_subindices: bool
) -> pd.DataFrame:
    """Build the 'Metric Descriptions' table (Column / Description / Unit).

    Score ranges and formulas reflect the actual computation: each metric is
    ranked into quintiles, convex-mapped (1->0, 2->1, 3->2, 4->5, 5->10) and
    summed within each sub-index, then combined multiplicatively.
    """
    n_h = len(_HAZARD_CLIMATE_METRICS if climate_hazards_only else _HAZARD_METRICS)
    h_scope = "climate hazards only" if climate_hazards_only else "all hazards"
    if normalize_subindices:
        formula = "norm(H) x (norm(T) + norm(A))"
        formula_ext = "norm(H_extended) x (norm(T) + norm(A))"
        cc_range = "0–2"
    else:
        formula = "H x (T + A)"
        formula_ext = "H_extended x (T + A)"
        cc_range = "composite score"

    rows: list[tuple[str, str, str]] = []
    rows.append(("— Identifiers —", "", ""))
    rows += _DESC_IDENTIFIERS
    rows.append(("— Traffic —", "", ""))
    rows += _DESC_TRAFFIC
    rows.append(("— Hazard exposure metrics —", "", ""))
    rows += _DESC_HAZARD
    rows.append(("— Travel disruption metrics —", "", ""))
    rows += _DESC_TRAVEL
    rows.append(("— Local accessibility metrics —", "", ""))
    rows += _DESC_ACCESS
    rows.append(("— Per-metric derived columns —", "", ""))
    rows += [
        ("<metric>_norm", "Min-max normalised metric value (log-transformed first for skewed metrics)", "0–1"),
        ("<metric>_q", "Quintile score: 0 = no exposure, 1 = lowest 20% of exposed sections, 5 = highest 20%", "0–5"),
        ("<metric>_cv", "Convex weight from the quintile score (1→0, 2→1, 3→2, 4→5, 5→10)", "0–10"),
    ]
    rows.append(("— Sub-indices and combined criticality —", "", ""))
    rows += [
        ("H", f"Hazard-exposure sub-index ({h_scope}): sum of convex scores of the {n_h} hazard metrics", f"0–{n_h * 10}"),
        ("H_extended", "Extended hazard-exposure sub-index (all 7 hazards): sum of convex scores", "0–70"),
        ("T", "National-scale travel-disruption sub-index: sum of convex scores of the 4 disruption metrics", "0–40"),
        ("A", "Local-accessibility sub-index: sum of convex scores of the 7 accessibility metrics", "0–70"),
        ("climate_criticality", f"Combined criticality ({h_scope}): {formula}", cc_range),
        ("climate_criticality_extended", f"Combined criticality (all 7 hazards): {formula_ext}", cc_range),
        ("H_class / T_class / A_class", "Sub-index criticality class (quintiles; 0 = No criticality)", "No criticality – Very High"),
        ("H_class_extended", "All-hazards hazard-exposure criticality class", "No criticality – Very High"),
        ("climate_criticality_class", f"Combined criticality class ({h_scope})", "No criticality – Very High"),
        ("climate_criticality_class_extended", "Combined criticality class (all 7 hazards)", "No criticality – Very High"),
    ]
    return pd.DataFrame(rows, columns=["Column", "Description", "Unit"])


def export_climate_criticality_excel(
    gdf: gpd.GeoDataFrame,
    output_path: Path,
    climate_hazards_only: bool = True,
    normalize_subindices: bool = True,
) -> None:
    """Write the multi-sheet, formatted climate-criticality workbook.

    Sheets (in order): Overview, Overview Extended, Hazard Exposure, Hazard
    Exposure Extended, National-Scale Disruption, Local Accessibility, Metric
    Descriptions. The plain sheets use the selected hazard set (climate-only
    when *climate_hazards_only*); the "Extended" sheets always use all 7 hazards
    (columns suffixed ``_extended``). Every data sheet is ordered from most to
    least critical (climate_criticality, then H, T, A).
    """
    df = pd.DataFrame(gdf.drop(columns=gdf.geometry.name))
    df = df.reset_index(drop=True)
    if "objectid" not in df.columns:
        df["objectid"] = range(1, len(df) + 1)

    # Order every sheet from most to least critical.
    sort_cols = [c for c in ["climate_criticality", "H", "T", "A"] if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols, ascending=False).reset_index(drop=True)

    hazard_metrics = _HAZARD_CLIMATE_METRICS if climate_hazards_only else _HAZARD_METRICS

    def metric_pairs(metrics):
        pairs = []
        for c in metrics:
            pairs.extend([c, f"{c}_q", f"{c}_cv"])
        return pairs

    overview = _EXCEL_ID_COLS + _EXCEL_TRAFFIC_COLS + [
        "H", "T", "A", "climate_criticality",
        "H_class", "T_class", "A_class", "climate_criticality_class",
    ]
    overview_ext = _EXCEL_ID_COLS + _EXCEL_TRAFFIC_COLS + [
        "H_extended", "T", "A", "climate_criticality_extended",
        "H_class_extended", "T_class", "A_class", "climate_criticality_class_extended",
    ]
    hazard = _EXCEL_ID_COLS + metric_pairs(hazard_metrics) + ["H", "H_class"]
    hazard_ext = _EXCEL_ID_COLS + metric_pairs(_HAZARD_METRICS) + ["H_extended", "H_class_extended"]
    travel = _EXCEL_ID_COLS + metric_pairs(_TRAVEL_METRICS) + ["T", "T_class"]
    access = _EXCEL_ID_COLS + metric_pairs(_ACCESSIBILITY_METRICS) + ["A", "A_class"]

    def existing(cols):
        return [c for c in cols if c in df.columns]

    descriptions = _metric_descriptions_df(climate_hazards_only, normalize_subindices)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df[existing(overview)].to_excel(writer, sheet_name="Overview", index=False)
        df[existing(overview_ext)].to_excel(writer, sheet_name="Overview Extended", index=False)
        df[existing(hazard)].to_excel(writer, sheet_name="Hazard Exposure", index=False)
        df[existing(hazard_ext)].to_excel(writer, sheet_name="Hazard Exposure Extended", index=False)
        df[existing(travel)].to_excel(writer, sheet_name="National-Scale Disruption", index=False)
        df[existing(access)].to_excel(writer, sheet_name="Local Accessibility", index=False)
        descriptions.to_excel(writer, sheet_name="Metric Descriptions", index=False)

    _format_climate_workbook(output_path)
    print(f"Saved climate-criticality workbook -> {output_path}")


def save_climate_criticality_geospatial(
    gdf: gpd.GeoDataFrame,
    parquet_path: Path,
) -> None:
    """Save the scored, deduplicated network as Parquet (working CRS) + mirrored Excel.

    No GeoPackage is written — the reprojected geospatial copy lives as the
    ``climate_criticality`` feature class in the results GDB, produced by
    ``plot_combined_climate_criticality``.
    """
    parquet_path = Path(parquet_path)
    parquet_path.parent.mkdir(parents=True, exist_ok=True)

    gdf.to_parquet(parquet_path)
    print(f"Saved climate criticality -> {parquet_path}")
    save_excel_mirror(gdf, parquet_path)
